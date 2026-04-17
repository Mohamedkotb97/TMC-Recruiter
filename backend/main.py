"""FastAPI backend — ingests LinkedIn captures, drafts AI replies, exports Excel."""

import os
import io
import csv
import json as json_lib
import threading
from concurrent.futures import ThreadPoolExecutor
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from typing import Optional, List
from fastapi import FastAPI, Depends, Header, HTTPException, Query, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_, func
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from models import (
    init_db, get_db, engine, SessionLocal,
    Candidate, Conversation, Message, ReplyDraft,
    Role, CandidateRoleMatch, MessageTemplate, ReplyClassification, Task,
    User, Session as SessionRow, TalentPool, Campaign, CampaignRecipient,
    StageHistory, Setting, candidate_pool_assoc,
)
import ai_service


def _jload(s: Optional[str], default):
    if not s:
        return default
    try:
        return json_lib.loads(s)
    except Exception:
        return default


def _jdump(v) -> Optional[str]:
    if v is None:
        return None
    return json_lib.dumps(v, ensure_ascii=False)


def _role_dict(r: Role) -> dict:
    return {
        "id": r.id,
        "title": r.title,
        "company": r.company or "",
        "location": r.location or "",
        "seniority": r.seniority or "",
        "employment_type": r.employment_type or "",
        "jd_text": r.jd_text or "",
        "must_have": _jload(r.must_have_json, []),
        "nice_to_have": _jload(r.nice_to_have_json, []),
        "target_companies": _jload(r.target_companies_json, []),
        "target_titles": _jload(r.target_titles_json, []),
        "search_keywords": _jload(r.search_keywords_json, []),
        "outreach_angle": r.outreach_angle or "",
        "persona": r.persona or "",
        "exclusions": r.exclusions or "",
        "status": r.status or "open",
        "created_at": r.created_at.isoformat() if r.created_at else "",
        "updated_at": r.updated_at.isoformat() if r.updated_at else "",
        "match_count": len(r.matches) if r.matches is not None else 0,
        "template_count": len(r.templates) if r.templates is not None else 0,
    }


def _candidate_summary(c: Candidate) -> dict:
    return {
        "id": c.id,
        "full_name": c.full_name,
        "current_title": c.current_title or "",
        "current_company": c.current_company or "",
        "headline": c.headline or "",
        "location": c.location or "",
        "about": c.about or "",
        "profile_url": c.profile_url or "",
        "stage": c.stage or "New",
        "email": c.email or "",
        "phone": c.phone or "",
        "skills": _jload(c.skills_json, []),
        "languages": _jload(c.languages_json, []),
        "visa_status": c.visa_status or "",
        "open_to_relocation": bool(c.open_to_relocation),
        "years_experience": c.years_experience,
        "salary_expectation": c.salary_expectation or "",
        "tags": [t.strip() for t in (c.tags or "").split(",") if t.strip()],
        "owner_id": c.owner_id,
        "source": c.source or "manual",
        "last_activity_at": c.last_activity_at.isoformat() if c.last_activity_at else "",
    }


def _user_dict(u, *, reveal_key: bool = False):
    if not u:
        return None
    d = {
        "id": u.id,
        "username": u.username,
        "display_name": u.display_name or u.username,
        "email": u.email or "",
        "role": u.role,
        "is_admin": bool(getattr(u, "is_admin", False)),
        "is_active": bool(u.is_active),
        "created_at": u.created_at.isoformat() if u.created_at else "",
        "has_api_key": bool(getattr(u, "api_key", None)),
    }
    if reveal_key:
        d["api_key"] = getattr(u, "api_key", None) or ""
    return d

# Note: require_admin is defined below, after current_user.


# ========== Settings (API keys stored in DB) ==========

def get_setting(db: Session, key: str, default: str = "") -> str:
    row = db.query(Setting).filter(Setting.key == key).first()
    return (row.value if row and row.value is not None else default)


def set_setting(db: Session, key: str, value: str):
    row = db.query(Setting).filter(Setting.key == key).first()
    if row:
        row.value = value
    else:
        db.add(Setting(key=key, value=value))
    db.commit()


# ========== Auth ==========
#
# Two ways to authenticate:
#   1. X-Session-Token  — set by the dashboard after username+password login.
#   2. X-API-Key        — a PER-USER key that the recruiter pastes into the
#                         Chrome extension so every saved conversation is
#                         tagged with `owner_user_id`. Generated on signup,
#                         rotatable from Settings.
#
# There is a global bootstrap key `CRM_API_KEY` (env var) that ONLY counts as
# an "admin / backend-owner" fallback — it never sets owner_user_id. Dashboard
# users MUST log in; the global key alone will no longer let you into the UI.

import secrets as _secrets

API_KEY = os.environ.get("CRM_API_KEY", "dev-key-change-me")


def _new_user_api_key() -> str:
    """32-byte URL-safe token, prefixed so it's recognisable in logs."""
    return "tmc_" + _secrets.token_urlsafe(24)


def _session_user(db: Session, token: Optional[str]) -> Optional[User]:
    if not token:
        return None
    s = db.query(SessionRow).filter(SessionRow.token == token).first()
    if not s:
        return None
    if s.expires_at and s.expires_at < datetime.utcnow():
        db.delete(s); db.commit()
        return None
    return s.user


def _api_key_user(db: Session, key: Optional[str]) -> Optional[User]:
    if not key:
        return None
    u = db.query(User).filter(User.api_key == key, User.is_active == True).first()
    return u


def require_api_key(
    x_api_key: str = Header(None, alias="X-API-Key"),
    x_session_token: str = Header(None, alias="X-Session-Token"),
    db: Session = Depends(get_db),
):
    """Accept a PER-USER api key, a dashboard session token, or the global
    bootstrap key (admin fallback only — no user context)."""
    # 1. Per-user API key (extension flow)
    if x_api_key and _api_key_user(db, x_api_key):
        return True
    # 2. Dashboard session
    u = _session_user(db, x_session_token)
    if u and u.is_active:
        return True
    # 3. Global bootstrap key — kept only for localhost/admin tooling
    if x_api_key and x_api_key == API_KEY:
        return True
    raise HTTPException(status_code=401, detail="Invalid API key or session token")


def require_user(
    user: Optional[User] = Depends(current_user),
) -> User:
    """Endpoints that MUST be attributed to a user (bulk ingest, save thread).
    The global CRM_API_KEY is rejected here on purpose — conversations saved
    without an owner end up invisible in every non-admin dashboard.
    """
    if not user:
        raise HTTPException(
            status_code=401,
            detail=(
                "This endpoint needs a personal extension key. "
                "Open the dashboard → Settings → Chrome Extension and copy your "
                "tmc_... key into the extension popup. The shared CRM_API_KEY is "
                "for admin tooling only."
            ),
        )
    return user


def current_user(
    x_session_token: str = Header(None, alias="X-Session-Token"),
    x_api_key: str = Header(None, alias="X-API-Key"),
    db: Session = Depends(get_db),
) -> Optional[User]:
    """Who is the current user?
    Preference order: session token → per-user API key → None (global bootstrap key)."""
    u = _session_user(db, x_session_token)
    if u:
        return u
    return _api_key_user(db, x_api_key)


def require_admin(
    user: Optional[User] = Depends(current_user),
    x_api_key: str = Header(None, alias="X-API-Key"),
) -> Optional[User]:
    """Admin-only. The global `CRM_API_KEY` still works (backend owner)."""
    if x_api_key and x_api_key == API_KEY:
        return None
    if not user or not getattr(user, "is_admin", False):
        raise HTTPException(403, "Admin only")
    return user


def _ensure_default_user(db: Session):
    """Seed a default admin user on first boot so the dashboard can be used immediately."""
    if db.query(User).count() == 0:
        default_pw = os.environ.get("DEFAULT_ADMIN_PASSWORD", "tmc-admin")
        db.add(User(
            username="habib",
            password_hash=ai_service.hash_password(default_pw),
            display_name="Habib Touil",
            role="admin",
            is_active=True,
            is_admin=True,
            api_key=_new_user_api_key(),
        ))
        db.commit()
        print(f"[info] seeded default user 'habib' with password '{default_pw}' — CHANGE IT.")
    else:
        # Make sure at least one user has is_admin=True (back-compat for existing DBs)
        if db.query(User).filter(User.is_admin == True).count() == 0:
            first = db.query(User).filter(User.role == "admin").first() or db.query(User).first()
            if first:
                first.is_admin = True
                db.commit()
    # Backfill API keys for any user that doesn't have one yet.
    for u in db.query(User).filter((User.api_key.is_(None)) | (User.api_key == "")).all():
        u.api_key = _new_user_api_key()
    db.commit()

    # Backfill orphan ownership — any candidate with owner_user_id=NULL that
    # has at least one conversation owned by a user gets that user as owner.
    # Prevents older data from being invisible after the multi-user upgrade.
    try:
        orphans = db.query(Candidate).filter(Candidate.owner_user_id.is_(None)).all()
        patched = 0
        for cand in orphans:
            conv = next(
                (c for c in cand.conversations if c.owner_user_id),
                None,
            )
            if conv:
                cand.owner_user_id = conv.owner_user_id
                patched += 1
        if patched:
            db.commit()
            print(f"[info] backfilled owner_user_id on {patched} orphan candidate(s)")
    except Exception as e:
        print(f"[warn] orphan backfill skipped: {e}")


# ========== Background analysis worker ==========
#
# When the Chrome extension bulk-uploads conversations we want to immediately
# kick off Apify enrichment + Claude analysis WITHOUT blocking the upload
# response (the extension might be syncing 100+ threads at a time).
#
# We use a small ThreadPoolExecutor. Each job opens its own DB session,
# runs the enrichment + analysis, updates `analysis_status`, and commits.
# Errors are caught and stored in `analysis_error` so the UI can surface them.

# DISABLE_BG_ANALYSIS=1 turns the worker off (for serverless hosts where the
# process dies between requests). Conversations still save fine — the brief
# fills in the next time the recruiter opens the thread (on-demand analyze).
_BG_ANALYSIS_ENABLED = os.environ.get("DISABLE_BG_ANALYSIS", "0") not in ("1", "true", "yes")
_ANALYSIS_POOL = (
    ThreadPoolExecutor(max_workers=3, thread_name_prefix="conv-analyze")
    if _BG_ANALYSIS_ENABLED else None
)
_ANALYSIS_LOCK = threading.Lock()
_ANALYSIS_INFLIGHT: set[int] = set()  # conversation ids currently being processed


def _run_conv_analysis_job(conv_id: int, do_enrich: bool = True) -> None:
    """Thread-pool job: enrich via Apify (if possible) + run Claude analysis.

    Safe to call repeatedly — noop if already done unless `force` is used in
    the sync endpoint.
    """
    # Prevent duplicate jobs for the same conversation
    with _ANALYSIS_LOCK:
        if conv_id in _ANALYSIS_INFLIGHT:
            return
        _ANALYSIS_INFLIGHT.add(conv_id)

    db = SessionLocal()
    try:
        conv = db.query(Conversation).filter(Conversation.id == conv_id).first()
        if not conv:
            return
        conv.analysis_status = "analyzing"
        conv.analysis_error = None
        db.commit()

        cand = conv.candidate
        if not cand:
            conv.analysis_status = "failed"
            conv.analysis_error = "candidate missing"
            db.commit()
            return

        # --- optional Apify enrichment ---
        if do_enrich and (not cand.about or not cand.skills_json):
            apify_key = get_setting(db, "apify_api_key", "") or get_setting(db, "proxycurl_api_key", "")
            profile_url = (cand.profile_url or "").strip()
            if apify_key and profile_url and profile_url.startswith("http"):
                actor_id = get_setting(db, "apify_actor_id", "") or ai_service.APIFY_DEFAULT_ACTOR
                try:
                    data = ai_service.apify_enrich_profile(profile_url, apify_key, actor_id=actor_id)
                    upsert_candidate(
                        db,
                        profile_url=profile_url,
                        full_name=data.get("full_name") or cand.full_name,
                        headline=data.get("headline"),
                        current_title=data.get("current_title"),
                        current_company=data.get("current_company"),
                        location=data.get("location"),
                        about=data.get("about"),
                        skills=data.get("skills"),
                        languages=data.get("languages"),
                        email=data.get("email"),
                        phone=data.get("phone"),
                        source="apify",
                    )
                    db.commit()
                except Exception as e:
                    print(f"[bg-analyze] apify failed for conv {conv_id}: {e}")

        cand = db.query(Candidate).filter(Candidate.id == cand.id).first()

        candidate_dict = {
            "full_name": cand.full_name,
            "headline": cand.headline,
            "current_title": cand.current_title,
            "current_company": cand.current_company,
            "location": cand.location,
            "about": cand.about,
            "skills": _jload(cand.skills_json, []),
        }

        pools = db.query(TalentPool).order_by(TalentPool.updated_at.desc()).limit(30).all()
        pools_dicts = [{"id": p.id, "name": p.name, "description": p.description or ""} for p in pools]

        roles = db.query(Role).filter(Role.status == "open").order_by(Role.updated_at.desc()).limit(30).all()
        roles_dicts = [
            {
                "id": r.id,
                "title": r.title,
                "company": r.company or "",
                "persona": r.persona or "",
                "must_have": _jload(r.must_have_json, []),
            }
            for r in roles
        ]

        result = ai_service.analyze_conversation_for_recruiter(
            candidate_dict,
            _build_msgs_for_ai(conv),
            pools_dicts,
            roles_dicts,
        )

        conv.summary = (result.get("conversation_brief") or "").strip() or conv.summary
        conv.sentiment = (result.get("sentiment") or conv.sentiment or "neutral")
        conv.person_brief = (result.get("person_brief") or "").strip()
        sugg_pool = result.get("suggested_pool_id")
        sugg_role = result.get("suggested_role_id")
        conv.suggested_pool_id = int(sugg_pool) if isinstance(sugg_pool, int) else None
        conv.suggested_role_id = int(sugg_role) if isinstance(sugg_role, int) else None
        conv.analyzed_at = datetime.utcnow()
        conv.analysis_status = "done"
        conv.analysis_error = None
        db.commit()
    except Exception as e:
        print(f"[bg-analyze] conv {conv_id} FAILED: {e}")
        try:
            conv = db.query(Conversation).filter(Conversation.id == conv_id).first()
            if conv:
                conv.analysis_status = "failed"
                conv.analysis_error = str(e)[:500]
                db.commit()
        except Exception:
            pass
    finally:
        db.close()
        with _ANALYSIS_LOCK:
            _ANALYSIS_INFLIGHT.discard(conv_id)


def enqueue_conv_analysis(conv_ids: list[int], do_enrich: bool = True) -> None:
    """Submit a batch of conversation IDs for background analysis.
    Safe to call with an empty list. Deduplicates in-flight jobs.
    No-op when DISABLE_BG_ANALYSIS=1 (serverless deployments); the dashboard's
    "↻ Brief pending" button then drives analysis on demand.
    """
    if not _ANALYSIS_POOL:
        return
    for cid in conv_ids:
        if not cid:
            continue
        _ANALYSIS_POOL.submit(_run_conv_analysis_job, cid, do_enrich)


# ========== App setup ==========

@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    _db = SessionLocal()
    try:
        _ensure_default_user(_db)
    finally:
        _db.close()
    yield


app = FastAPI(title="Recruiter CRM", version="1.0.0", lifespan=lifespan)

# CORS — Chrome extension posts from chrome-extension://<id> (always allowed
# via wildcard because the id changes across installs/devices). For the
# dashboard + any custom origins, set ALLOWED_ORIGINS as a comma-separated
# list in the env (e.g. "https://recruiter.tmc.com,https://staging.tmc.com").
# Default stays "*" for easy local dev.
_allowed = os.environ.get("ALLOWED_ORIGINS", "*").strip()
if _allowed == "*" or not _allowed:
    _cors_origins = ["*"]
    _cors_regex = None
else:
    _cors_origins = [o.strip() for o in _allowed.split(",") if o.strip()]
    # Always allow any chrome-extension origin so the popup + content script
    # can hit the API regardless of install id.
    _cors_regex = r"^chrome-extension://.*$"

app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_origin_regex=_cors_regex,
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ========== Pydantic schemas ==========

class MessageIn(BaseModel):
    sender: str
    timestamp: Optional[str] = ""
    body: str


class ConversationIn(BaseModel):
    candidate_name: str
    profile_url: str
    captured_at: str
    messages: list[MessageIn]
    source: str = "linkedin_messaging"
    thread_url: Optional[str] = ""   # unique LinkedIn thread URL, used for dedup
    # Optional profile enrichment lifted from the messaging header.
    headline: Optional[str] = ""
    current_title: Optional[str] = ""
    current_company: Optional[str] = ""
    location: Optional[str] = ""


class CandidateIn(BaseModel):
    full_name: str
    profile_url: str
    headline: Optional[str] = ""
    location: Optional[str] = ""
    current_title: Optional[str] = ""
    current_company: Optional[str] = ""
    about: Optional[str] = ""
    captured_at: Optional[str] = ""
    source: Optional[str] = "linkedin_profile"


class DraftRequest(BaseModel):
    tone: str = "professional"
    goal: str = "continue the conversation and move toward a call"


class DraftApproval(BaseModel):
    approved: bool


class NotesUpdate(BaseModel):
    notes: Optional[str] = None
    tags: Optional[list] = None
    stage: Optional[str] = None
    full_name: Optional[str] = None
    headline: Optional[str] = None
    current_title: Optional[str] = None
    current_company: Optional[str] = None
    location: Optional[str] = None
    about: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    skills: Optional[list] = None
    languages: Optional[list] = None
    visa_status: Optional[str] = None
    open_to_relocation: Optional[bool] = None
    years_experience: Optional[int] = None
    salary_expectation: Optional[str] = None
    owner_id: Optional[int] = None
    pool_ids: Optional[List[int]] = None


# ========== Helpers ==========

def _touch_activity(c: Candidate):
    c.last_activity_at = datetime.utcnow()


def _parse_msg_ts(raw: Optional[str]):
    """Parse the message 'timestamp_raw' (ISO 8601 from the Chrome extension) to a datetime.
    Returns None when parsing fails or input is empty — callers should fall back to created_at.
    """
    if not raw:
        return None
    s = str(raw).strip()
    if not s:
        return None
    try:
        # Python's fromisoformat handles 'Z' only in 3.11+; be defensive
        if s.endswith("Z"):
            s = s[:-1] + "+00:00"
        dt = datetime.fromisoformat(s)
        # Normalize to naive UTC for consistent comparison with other datetimes in DB
        if dt.tzinfo is not None:
            try:
                import datetime as _dt
                dt = dt.astimezone(_dt.timezone.utc).replace(tzinfo=None)
            except Exception:
                dt = dt.replace(tzinfo=None)
        return dt
    except Exception:
        return None


def _record_stage_change(db: Session, candidate_id: int, from_stage: str, to_stage: str,
                         user: Optional[User]):
    if from_stage == to_stage:
        return
    db.add(StageHistory(
        candidate_id=candidate_id,
        from_stage=from_stage or "",
        to_stage=to_stage or "",
        changed_by=user.id if user else None,
    ))


# JSON-list fields on Candidate we accept as either list OR comma-string.
_CAND_JSON_FIELDS = {"skills": "skills_json", "languages": "languages_json"}
_CAND_SIMPLE = {
    "full_name", "profile_url", "headline", "current_title", "current_company",
    "location", "about", "email", "phone", "visa_status", "years_experience",
    "salary_expectation", "tags", "notes", "source", "owner_id",
}


def _coerce_list(v):
    if v is None:
        return None
    if isinstance(v, list):
        return v
    if isinstance(v, str):
        return [x.strip() for x in v.split(",") if x.strip()]
    return None


def upsert_candidate(db: Session, profile_url: str, full_name: str, **extra) -> Candidate:
    """Find by profile URL (or email if url missing) or create. Update with any new extra fields.

    Also: auto-dedups on (linkedin_url, email) to prevent double entries,
    records last_activity, and never overwrites a non-empty field with a blank.
    """
    cand = None
    if profile_url:
        cand = db.query(Candidate).filter(Candidate.profile_url == profile_url).first()
    if not cand and extra.get("email"):
        cand = db.query(Candidate).filter(Candidate.email == extra["email"]).first()

    if cand is None:
        cand = Candidate(profile_url=profile_url or f"manual:{full_name}", full_name=full_name)
        db.add(cand)
        db.flush()
    else:
        if full_name:
            cand.full_name = full_name

    # Simple fields: only overwrite if new value is truthy.
    for k, v in extra.items():
        if k in _CAND_JSON_FIELDS:
            lst = _coerce_list(v)
            if lst is not None:
                setattr(cand, _CAND_JSON_FIELDS[k], _jdump(lst))
        elif k == "open_to_relocation":
            if v is not None:
                cand.open_to_relocation = bool(v)
        elif k in _CAND_SIMPLE:
            if v not in (None, ""):
                setattr(cand, k, v)

    _touch_activity(cand)
    return cand


# ========== Endpoints: capture ==========

@app.post("/api/candidates")
def save_candidate(
    payload: CandidateIn,
    db: Session = Depends(get_db),
    user: User = Depends(require_user),
):
    """Save/update a candidate profile captured from LinkedIn.
    Attributed to the saving user so it shows up in their dashboard.
    """
    cand = upsert_candidate(
        db,
        profile_url=payload.profile_url,
        full_name=payload.full_name,
        headline=payload.headline,
        location=payload.location,
        current_title=payload.current_title,
        current_company=payload.current_company,
        about=payload.about,
    )
    if not cand.owner_user_id:
        cand.owner_user_id = user.id
    db.commit()
    return {"id": cand.id, "full_name": cand.full_name}


@app.post("/api/conversations")
def save_conversation(
    payload: ConversationIn,
    db: Session = Depends(get_db),
    user: User = Depends(require_user),
):
    """Save a DM thread. Auto-creates candidate. Dedups by thread_url:
    re-saving the same thread replaces its messages (no duplicates).
    Requires a per-user personal key — shared/global keys are rejected so
    saved threads are always attributed to a dashboard user.
    """
    if not payload.messages:
        raise HTTPException(400, "No messages in payload")

    cand = upsert_candidate(
        db,
        profile_url=payload.profile_url or f"manual:{payload.candidate_name}",
        full_name=payload.candidate_name,
        headline=payload.headline,
        current_title=payload.current_title,
        current_company=payload.current_company,
        location=payload.location,
    )

    thread_url = (payload.thread_url or "").strip()

    existing = None
    if thread_url:
        existing = db.query(Conversation).filter(
            Conversation.thread_url == thread_url
        ).first()

    # Hash the new message set so we can detect "nothing changed" fast.
    new_msg_keys = [(m.sender or "", m.body or "") for m in payload.messages]

    if existing is not None:
        old_msg_keys = [(m.sender or "", m.body or "") for m in existing.messages]
        unchanged = old_msg_keys == new_msg_keys

        if unchanged:
            # Idempotent no-op — don't re-run Claude, don't rewrite messages.
            return {
                "conversation_id": existing.id,
                "candidate_id": cand.id,
                "message_count": len(existing.messages),
                "summary": existing.summary,
                "sentiment": existing.sentiment,
                "status": "unchanged",
            }

        # Thread grew / edited — replace messages atomically.
        for old in list(existing.messages):
            db.delete(old)
        db.flush()
        for m in payload.messages:
            db.add(Message(
                conversation_id=existing.id,
                sender=m.sender,
                body=m.body,
                timestamp_raw=m.timestamp,
            ))
        existing.captured_at = datetime.utcnow()
        existing.source = payload.source
        existing.analyzed_at = None
        existing.analysis_status = "pending"
        existing.analysis_error = None
        if not existing.owner_user_id:
            existing.owner_user_id = user.id
        db.commit()
        conv = existing
        status = "updated"
    else:
        conv = Conversation(
            candidate_id=cand.id,
            thread_url=thread_url or None,
            captured_at=datetime.utcnow(),
            source=payload.source,
            analysis_status="pending",
            owner_user_id=user.id,
        )
        db.add(conv)
        db.flush()
        for m in payload.messages:
            db.add(Message(
                conversation_id=conv.id,
                sender=m.sender,
                body=m.body,
                timestamp_raw=m.timestamp,
            ))
        db.commit()
        status = "created"

    # Tag candidate with owner if not yet set. (Shared profiles stay owned
    # by the original creator; the dashboard's per-conversation ownership
    # check still makes this visible to whoever saved the thread.)
    if not cand.owner_user_id:
        cand.owner_user_id = user.id
        db.commit()

    # Kick off background enrichment + AI analysis (Apify + Claude).
    # We do NOT block the save — the recruiter sees the conversation
    # immediately; the brief fills in when the worker finishes.
    enqueue_conv_analysis([conv.id], do_enrich=True)

    return {
        "conversation_id": conv.id,
        "candidate_id": cand.id,
        "message_count": len(payload.messages),
        "summary": conv.summary,
        "sentiment": conv.sentiment,
        "status": status,
        "analysis_status": "pending",
    }


# ========== Bulk conversations (fast batched ingest from Chrome extension) ==========

class BulkConversationsIn(BaseModel):
    """Accepts N conversations in one call. We upsert fast (no AI) and return per-item status.
    The extension can spam this with 100+ conversations without the backend doing any
    AI work — analysis happens later when the recruiter opens a conversation."""
    conversations: list[ConversationIn]
    source: Optional[str] = "linkedin_extension"


@app.post("/api/conversations/bulk")
def bulk_save_conversations(
    payload: BulkConversationsIn,
    db: Session = Depends(get_db),
    user: User = Depends(require_user),
):
    """Ingest many conversations at once. Deduplicates by thread_url — no duplicates
    on repeated sync. AI analysis is NOT run here to keep it fast; call
    /api/conversations/{id}/analyze when the recruiter opens a thread (or click
    'Analyze all new' to batch).
    """
    if not payload.conversations:
        raise HTTPException(400, "No conversations provided")

    created = 0
    updated = 0
    unchanged = 0
    skipped = 0
    conv_ids: list[int] = []
    enqueue_ids: list[int] = []  # new/updated — need background analysis

    for c in payload.conversations:
        try:
            if not c.messages:
                skipped += 1
                continue

            cand = upsert_candidate(
                db,
                profile_url=c.profile_url or f"manual:{c.candidate_name}",
                full_name=c.candidate_name,
                headline=c.headline,
                current_title=c.current_title,
                current_company=c.current_company,
                location=c.location,
                source=payload.source or c.source or "linkedin_extension",
            )
            if not cand.owner_user_id:
                cand.owner_user_id = user.id

            thread_url = (c.thread_url or "").strip()
            existing = None
            if thread_url:
                existing = db.query(Conversation).filter(
                    Conversation.thread_url == thread_url
                ).first()

            new_keys = [(m.sender or "", m.body or "") for m in c.messages]

            if existing is not None:
                old_keys = [(m.sender or "", m.body or "") for m in existing.messages]
                if old_keys == new_keys:
                    unchanged += 1
                    conv_ids.append(existing.id)
                    continue
                # Replace messages atomically
                for old in list(existing.messages):
                    db.delete(old)
                db.flush()
                for m in c.messages:
                    db.add(Message(
                        conversation_id=existing.id,
                        sender=m.sender,
                        body=m.body,
                        timestamp_raw=m.timestamp,
                    ))
                existing.captured_at = datetime.utcnow()
                existing.source = c.source or payload.source or existing.source
                # Invalidate prior analysis since the thread changed
                existing.analyzed_at = None
                existing.analysis_status = "pending"
                existing.analysis_error = None
                # Claim the thread if it had no owner (legacy rows from
                # before per-user scoping) so the saver can see it.
                if not existing.owner_user_id:
                    existing.owner_user_id = user.id
                updated += 1
                conv_ids.append(existing.id)
                enqueue_ids.append(existing.id)
            else:
                conv = Conversation(
                    candidate_id=cand.id,
                    thread_url=thread_url or None,
                    captured_at=datetime.utcnow(),
                    source=c.source or payload.source or "linkedin_extension",
                    analysis_status="pending",
                    owner_user_id=user.id,
                )
                db.add(conv)
                db.flush()
                for m in c.messages:
                    db.add(Message(
                        conversation_id=conv.id,
                        sender=m.sender,
                        body=m.body,
                        timestamp_raw=m.timestamp,
                    ))
                created += 1
                conv_ids.append(conv.id)
                enqueue_ids.append(conv.id)
        except Exception as e:
            print(f"[warn] bulk conv item failed: {e}")
            skipped += 1

    db.commit()

    # Kick off background enrichment + AI analysis for everything new/updated.
    # Response returns immediately — the extension can move on to the next batch.
    enqueue_conv_analysis(enqueue_ids, do_enrich=True)

    return {
        "total": len(payload.conversations),
        "created": created,
        "updated": updated,
        "unchanged": unchanged,
        "skipped": skipped,
        "conversation_ids": conv_ids,
        "queued_for_analysis": len(enqueue_ids),
    }


# ========== Per-conversation AI analysis (person brief + pool/role suggestion) ==========

def _build_msgs_for_ai(conv: Conversation) -> list[dict]:
    msgs = sorted(
        conv.messages,
        key=lambda m: _parse_msg_ts(m.timestamp_raw) or m.created_at or datetime.min,
    )
    # Keep roughly the most recent 40 messages — older stuff adds noise and token cost
    if len(msgs) > 40:
        msgs = msgs[-40:]
    return [{"sender": m.sender or "", "body": m.body or ""} for m in msgs]


@app.post("/api/conversations/{conversation_id}/analyze")
def analyze_conversation(
    conversation_id: int,
    enrich: bool = True,
    force: bool = False,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    """Generate conversation brief + person brief + pool & role suggestions for one conversation.

    Also (best-effort) enriches the candidate via Apify if we don't have an
    'about' / skills yet and an Apify key is configured.
    """
    conv = db.query(Conversation).filter(Conversation.id == conversation_id).first()
    if not conv:
        raise HTTPException(404, "Conversation not found")
    cand = conv.candidate
    if not cand:
        raise HTTPException(404, "Candidate missing for conversation")

    # Optional Apify enrichment — only if we don't already have rich profile data
    enriched_via_apify = False
    if enrich and (not cand.about or not cand.skills_json):
        apify_key = get_setting(db, "apify_api_key", "") or get_setting(db, "proxycurl_api_key", "")
        profile_url = (cand.profile_url or "").strip()
        if apify_key and profile_url and profile_url.startswith("http"):
            actor_id = get_setting(db, "apify_actor_id", "") or ai_service.APIFY_DEFAULT_ACTOR
            try:
                data = ai_service.apify_enrich_profile(profile_url, apify_key, actor_id=actor_id)
                upsert_candidate(
                    db,
                    profile_url=profile_url,
                    full_name=data.get("full_name") or cand.full_name,
                    headline=data.get("headline"),
                    current_title=data.get("current_title"),
                    current_company=data.get("current_company"),
                    location=data.get("location"),
                    about=data.get("about"),
                    skills=data.get("skills"),
                    languages=data.get("languages"),
                    email=data.get("email"),
                    phone=data.get("phone"),
                    source="apify",
                )
                db.commit()
                enriched_via_apify = True
            except Exception as e:
                print(f"[warn] analyze: apify enrichment failed: {e}")

    # Re-fetch candidate so we pass the freshest data into the AI prompt
    cand = db.query(Candidate).filter(Candidate.id == cand.id).first()

    if conv.analyzed_at and not force and conv.summary and conv.person_brief:
        # Cached — return what we already have without burning Claude tokens
        return _conv_analysis_payload(conv, cand, cached=True)

    candidate_dict = {
        "full_name": cand.full_name,
        "headline": cand.headline,
        "current_title": cand.current_title,
        "current_company": cand.current_company,
        "location": cand.location,
        "about": cand.about,
        "skills": _jload(cand.skills_json, []),
    }

    pools = db.query(TalentPool).order_by(TalentPool.updated_at.desc()).limit(30).all()
    pools_dicts = [{"id": p.id, "name": p.name, "description": p.description or ""} for p in pools]

    roles = db.query(Role).filter(Role.status == "open").order_by(Role.updated_at.desc()).limit(30).all()
    roles_dicts = [
        {
            "id": r.id,
            "title": r.title,
            "company": r.company or "",
            "persona": r.persona or "",
            "must_have": _jload(r.must_have_json, []),
        }
        for r in roles
    ]

    try:
        result = ai_service.analyze_conversation_for_recruiter(
            candidate_dict,
            _build_msgs_for_ai(conv),
            pools_dicts,
            roles_dicts,
        )
    except Exception as e:
        raise HTTPException(500, f"Analysis failed: {e}")

    conv.summary = (result.get("conversation_brief") or "").strip() or conv.summary
    conv.sentiment = (result.get("sentiment") or conv.sentiment or "neutral")
    conv.person_brief = (result.get("person_brief") or "").strip()
    sugg_pool = result.get("suggested_pool_id")
    sugg_role = result.get("suggested_role_id")
    conv.suggested_pool_id = int(sugg_pool) if isinstance(sugg_pool, int) else None
    conv.suggested_role_id = int(sugg_role) if isinstance(sugg_role, int) else None
    conv.analyzed_at = datetime.utcnow()
    conv.analysis_status = "done"
    conv.analysis_error = None
    db.commit()

    payload = _conv_analysis_payload(conv, cand, cached=False)
    payload["enriched_via_apify"] = enriched_via_apify
    payload["suggested_pool_reason"] = result.get("suggested_pool_reason", "")
    payload["suggested_role_reason"] = result.get("suggested_role_reason", "")
    payload["reply_hint"] = result.get("reply_hint", "")
    return payload


def _conv_analysis_payload(conv: Conversation, cand: Candidate, cached: bool) -> dict:
    return {
        "conversation_id": conv.id,
        "candidate_id": cand.id,
        "conversation_brief": conv.summary or "",
        "person_brief": conv.person_brief or "",
        "sentiment": conv.sentiment or "neutral",
        "suggested_pool_id": conv.suggested_pool_id,
        "suggested_role_id": conv.suggested_role_id,
        "pool_id": conv.pool_id,
        "role_id": conv.role_id,
        "analyzed_at": conv.analyzed_at.isoformat() if conv.analyzed_at else "",
        "analysis_status": getattr(conv, "analysis_status", None) or "done",
        "cached": cached,
    }


@app.get("/api/analysis/status")
def analysis_status(
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    """Returns a summary of the background-analysis queue so the UI can show
    'Analyzing N of M' while the Chrome extension sync finishes."""
    counts: dict[str, int] = {}
    rows = db.query(Conversation.analysis_status, func.count(Conversation.id)).group_by(Conversation.analysis_status).all()
    for status, n in rows:
        counts[status or "pending"] = int(n)
    return {
        "pending": counts.get("pending", 0),
        "analyzing": counts.get("analyzing", 0),
        "done": counts.get("done", 0),
        "failed": counts.get("failed", 0),
        "inflight": len(_ANALYSIS_INFLIGHT),
    }


@app.post("/api/analysis/rerun_pending")
def rerun_pending_analysis(
    limit: int = 200,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    """Manually enqueue every conversation whose status is pending / failed.
    Useful after adding an Apify key or if the server restarted mid-sync."""
    rows = (
        db.query(Conversation.id)
        .filter(Conversation.analysis_status.in_(("pending", "failed")))
        .order_by(Conversation.captured_at.desc())
        .limit(limit)
        .all()
    )
    ids = [r[0] for r in rows]
    enqueue_conv_analysis(ids, do_enrich=True)
    return {"queued": len(ids)}


class ConvMetaIn(BaseModel):
    pool_id: Optional[int] = None
    role_id: Optional[int] = None
    clear_pool: bool = False
    clear_role: bool = False


@app.patch("/api/conversations/{conversation_id}/meta")
def update_conversation_meta(
    conversation_id: int,
    payload: ConvMetaIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    """Recruiter's pick of pool + role for this conversation (overrides AI suggestion)."""
    conv = db.query(Conversation).filter(Conversation.id == conversation_id).first()
    if not conv:
        raise HTTPException(404, "Conversation not found")

    if payload.clear_pool:
        conv.pool_id = None
    elif payload.pool_id is not None:
        p = db.query(TalentPool).filter(TalentPool.id == payload.pool_id).first()
        if not p:
            raise HTTPException(404, "Pool not found")
        conv.pool_id = p.id
        # Also add the candidate to the pool so it's actually in that talent list
        if conv.candidate and conv.candidate not in p.candidates:
            p.candidates.append(conv.candidate)

    if payload.clear_role:
        conv.role_id = None
    elif payload.role_id is not None:
        r = db.query(Role).filter(Role.id == payload.role_id).first()
        if not r:
            raise HTTPException(404, "Role not found")
        conv.role_id = r.id

    db.commit()
    return {"ok": True, "pool_id": conv.pool_id, "role_id": conv.role_id}


# ========== Kanban membership (manual add/remove) ==========

class KanbanToggleIn(BaseModel):
    in_kanban: bool


@app.patch("/api/candidates/{candidate_id}/kanban")
def toggle_kanban(
    candidate_id: int,
    payload: KanbanToggleIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    """Add or remove a candidate from the Kanban board. The board only shows
    people the recruiter has explicitly added."""
    c = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not c:
        raise HTTPException(404, "Candidate not found")
    c.in_kanban = bool(payload.in_kanban)
    db.commit()
    return {"ok": True, "in_kanban": c.in_kanban}


# ========== Endpoints: list & retrieve ==========

@app.get("/api/candidates")
def list_candidates(
    stage: Optional[str] = None,
    q: Optional[str] = None,
    location: Optional[str] = None,
    skill: Optional[str] = None,
    pool_id: Optional[int] = None,
    owner_id: Optional[int] = None,
    as_user_id: Optional[int] = None,
    tag: Optional[str] = None,
    visa: Optional[str] = None,
    min_years: Optional[int] = None,
    open_to_relocation: Optional[bool] = None,
    db: Session = Depends(get_db),
    user: Optional[User] = Depends(current_user),
    _: bool = Depends(require_api_key),
):
    query = db.query(Candidate)
    # Admin impersonation: if an admin passes ?as_user_id=N, scope the query to
    # that user's data exactly as if that user had made the request. Non-admins
    # trying to pass this param are ignored.
    scope_user_id: Optional[int] = None
    if user and getattr(user, "is_admin", False) and as_user_id is not None:
        scope_user_id = int(as_user_id)
    elif user and not getattr(user, "is_admin", False):
        scope_user_id = user.id

    # Multi-user scoping: non-admins see candidates they own OR candidates
    # that have at least one conversation they own. (Previously we filtered
    # ONLY on Candidate.owner_user_id, which hid conversations the user just
    # saved whenever the candidate row was first created by someone else —
    # the classic "I saved 5 threads but the Inbox is empty" bug.)
    if scope_user_id is not None:
        owned_candidate_ids = db.query(Conversation.candidate_id).filter(
            Conversation.owner_user_id == scope_user_id
        ).distinct().subquery()
        query = query.filter(or_(
            Candidate.owner_user_id == scope_user_id,
            Candidate.id.in_(owned_candidate_ids),
        ))
    if stage:
        query = query.filter(Candidate.stage == stage)
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Candidate.full_name.ilike(like),
            Candidate.current_company.ilike(like),
            Candidate.current_title.ilike(like),
            Candidate.headline.ilike(like),
            Candidate.email.ilike(like),
        ))
    if location:
        query = query.filter(Candidate.location.ilike(f"%{location}%"))
    if skill:
        # skills_json is a JSON array text — case-insensitive substring match is fine for SQLite.
        query = query.filter(Candidate.skills_json.ilike(f"%{skill}%"))
    if tag:
        query = query.filter(Candidate.tags.ilike(f"%{tag}%"))
    if visa:
        query = query.filter(Candidate.visa_status.ilike(f"%{visa}%"))
    if min_years is not None:
        query = query.filter(Candidate.years_experience >= min_years)
    if open_to_relocation is not None:
        query = query.filter(Candidate.open_to_relocation == bool(open_to_relocation))
    if owner_id is not None:
        query = query.filter(Candidate.owner_id == owner_id)
    if pool_id is not None:
        query = query.join(candidate_pool_assoc, candidate_pool_assoc.c.candidate_id == Candidate.id)\
                     .filter(candidate_pool_assoc.c.pool_id == pool_id)

    rows = query.all()
    out = []
    for c in rows:
        # Find the latest message across all of this candidate's conversations
        last_msg = None
        last_ts = None
        last_conv_sentiment = ""
        last_conv_id = None
        for conv in c.conversations:
            for m in conv.messages:
                ts = _parse_msg_ts(m.timestamp_raw) or m.created_at or conv.captured_at
                if ts is None:
                    continue
                if last_ts is None or ts > last_ts:
                    last_ts = ts
                    last_msg = m
                    last_conv_sentiment = conv.sentiment or ""
                    last_conv_id = conv.id
        if last_msg is not None:
            body = last_msg.body or ""
            preview = body.replace("\r\n", " ").replace("\n", " ").strip()
            if len(preview) > 140:
                preview = preview[:140] + "..."
            last_message_from_me = (last_msg.sender or "").strip().lower() == "you"
            last_message_preview = preview
            last_message_at_iso = last_ts.isoformat() if last_ts else ""
        else:
            last_message_from_me = False
            last_message_preview = ""
            last_message_at_iso = c.last_activity_at.isoformat() if c.last_activity_at else ""

        out.append({
            "id": c.id,
            "full_name": c.full_name,
            "headline": c.headline,
            "current_company": c.current_company,
            "current_title": c.current_title,
            "location": c.location,
            "stage": c.stage,
            "tags": [t.strip() for t in (c.tags or "").split(",") if t.strip()],
            "profile_url": c.profile_url,
            "email": c.email or "",
            "skills": _jload(c.skills_json, [])[:6],
            "years_experience": c.years_experience,
            "conversation_count": len(c.conversations),
            "pool_ids": [p.id for p in c.pools],
            "owner_id": c.owner_id,
            "source": c.source or "manual",
            "last_activity_at": c.last_activity_at.isoformat() if c.last_activity_at else "",
            "updated_at": c.updated_at.isoformat() if c.updated_at else "",
            # Inbox fields
            "last_message_preview": last_message_preview,
            "last_message_at": last_message_at_iso,
            "last_message_from_me": last_message_from_me,
            "last_sentiment": last_conv_sentiment,
            "last_conversation_id": last_conv_id,
            "in_kanban": bool(getattr(c, "in_kanban", False)),
        })
    # Sort: most recent activity first (inbox-style).
    def _sort_key(r):
        return r.get("last_message_at") or r.get("last_activity_at") or r.get("updated_at") or ""
    out.sort(key=_sort_key, reverse=True)
    return out


@app.get("/api/stats")
def stats(
    as_user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    user: Optional[User] = Depends(current_user),
    _: bool = Depends(require_api_key),
):
    """Lightweight dashboard KPIs — stage counts + conversation/reply stats.
    Non-admins only see counts for THEIR own data (multi-user isolation).
    Admins may pass ?as_user_id=N to view stats as that user."""
    scope_user_id: Optional[int] = None
    if user and getattr(user, "is_admin", False) and as_user_id is not None:
        scope_user_id = int(as_user_id)
    elif user and not getattr(user, "is_admin", False):
        scope_user_id = user.id

    cand_q = db.query(Candidate)
    conv_q = db.query(Conversation)
    if scope_user_id is not None:
        # Same rule as list_candidates: a candidate is "yours" if you own
        # the row OR you own at least one conversation on it.
        owned_cand_ids = db.query(Conversation.candidate_id).filter(
            Conversation.owner_user_id == scope_user_id
        ).distinct().subquery()
        cand_q = cand_q.filter(or_(
            Candidate.owner_user_id == scope_user_id,
            Candidate.id.in_(owned_cand_ids),
        ))
        conv_q = conv_q.filter(Conversation.owner_user_id == scope_user_id)
    all_cands = cand_q.all()
    stage_counts: dict[str, int] = {}
    for c in all_cands:
        s = c.stage or "New"
        stage_counts[s] = stage_counts.get(s, 0) + 1

    convs = conv_q.all()
    total_convs = len(convs)
    # "Reply rate" proxy: conversations where at least one message is NOT from "You".
    with_reply = 0
    sentiment_counts = {"interested": 0, "maybe": 0, "not_interested": 0, "neutral": 0}
    for conv in convs:
        if any((m.sender or "").strip().lower() != "you" for m in conv.messages):
            with_reply += 1
        s = (conv.sentiment or "neutral").lower()
        sentiment_counts[s] = sentiment_counts.get(s, 0) + 1

    reply_rate = round((with_reply / total_convs) * 100) if total_convs else 0

    # Pending follow-ups: candidates in "Engaged"/"Sent"/"Replied" without a recent conversation (>5 days).
    from datetime import timedelta
    now = datetime.utcnow()
    stale_threshold = now - timedelta(days=5)
    pending_followups = 0
    for c in all_cands:
        if (c.stage or "") in ("Engaged", "Sent", "Replied", "Follow-up Needed"):
            if not c.conversations or all(
                (conv.captured_at or now) < stale_threshold for conv in c.conversations
            ):
                pending_followups += 1

    open_roles = db.query(Role).filter(Role.status == "open").count()
    pending_triage = db.query(ReplyClassification).filter(
        ReplyClassification.handled.is_(False)
    ).count()
    pending_tasks = db.query(Task).filter(Task.status == "pending").count()
    overdue_tasks = db.query(Task).filter(
        Task.status == "pending",
        Task.due_at.isnot(None),
        Task.due_at < now,
    ).count()

    return {
        "total_candidates": len(all_cands),
        "total_conversations": total_convs,
        "reply_rate_pct": reply_rate,
        "pending_followups": pending_followups,
        "stage_counts": stage_counts,
        "sentiment_counts": sentiment_counts,
        "open_roles": open_roles,
        "pending_triage": pending_triage,
        "pending_tasks": pending_tasks,
        "overdue_tasks": overdue_tasks,
    }


def _check_owns_candidate(user: Optional[User], c: Candidate):
    """Raise 404 if the logged-in (non-admin) user doesn't own this candidate.
    A user "owns" a candidate when:
      - Candidate.owner_user_id == user.id, OR
      - any Conversation on the candidate is owned by user.id.
    This keeps shared LinkedIn profiles (already created by another teammate)
    accessible to anyone who later captures a thread with them.
    """
    if user and not getattr(user, "is_admin", False):
        if c.owner_user_id and c.owner_user_id == user.id:
            return
        if any(conv.owner_user_id == user.id for conv in (c.conversations or [])):
            return
        if c.owner_user_id:
            raise HTTPException(404, "Not found")
        # Untagged (NULL owner) candidates stay visible — legacy rows from
        # before multi-user support. Admins can reassign from the Admin page.


def _check_owns_conversation(user: Optional[User], conv: Conversation):
    if user and not getattr(user, "is_admin", False):
        if conv.owner_user_id and conv.owner_user_id != user.id:
            raise HTTPException(404, "Not found")


@app.get("/api/candidates/{candidate_id}")
def get_candidate(
    candidate_id: int,
    db: Session = Depends(get_db),
    user: Optional[User] = Depends(current_user),
    _: bool = Depends(require_api_key),
):
    c = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not c:
        raise HTTPException(404, "Candidate not found")
    _check_owns_candidate(user, c)

    # Per-role matches + role info for this candidate
    matches = db.query(CandidateRoleMatch).filter(
        CandidateRoleMatch.candidate_id == c.id
    ).all()
    match_rows = []
    for m in matches:
        r = m.role
        match_rows.append({
            "id": m.id,
            "role_id": m.role_id,
            "role_title": r.title if r else "",
            "role_company": (r.company if r else "") or "",
            "fit_score": m.fit_score,
            "fit_reason": m.fit_reason or "",
            "fit_bullets": _jload(m.fit_bullets_json, []),
            "risk_flags": _jload(m.risk_flags_json, []),
            "stage": m.stage or "New",
        })

    # Pending tasks for this candidate
    tasks = db.query(Task).filter(
        Task.entity_type == "candidate",
        Task.entity_id == c.id,
        Task.status == "pending",
    ).order_by(Task.due_at.asc()).all()

    pools = [{"id": p.id, "name": p.name, "color": p.color} for p in c.pools]
    history = db.query(StageHistory).filter(
        StageHistory.candidate_id == c.id
    ).order_by(StageHistory.changed_at.desc()).limit(20).all()
    stage_history = [
        {
            "from": h.from_stage,
            "to": h.to_stage,
            "at": h.changed_at.isoformat() if h.changed_at else "",
            "changed_by": h.changed_by,
        }
        for h in history
    ]

    return {
        "id": c.id,
        "full_name": c.full_name,
        "headline": c.headline,
        "current_company": c.current_company,
        "current_title": c.current_title,
        "location": c.location,
        "about": c.about,
        "stage": c.stage,
        "tags": [t.strip() for t in (c.tags or "").split(",") if t.strip()],
        "notes": c.notes,
        "profile_url": c.profile_url,
        "email": c.email or "",
        "phone": c.phone or "",
        "skills": _jload(c.skills_json, []),
        "languages": _jload(c.languages_json, []),
        "visa_status": c.visa_status or "",
        "open_to_relocation": bool(c.open_to_relocation),
        "years_experience": c.years_experience,
        "salary_expectation": c.salary_expectation or "",
        "source": c.source or "manual",
        "owner_id": c.owner_id,
        "last_activity_at": c.last_activity_at.isoformat() if c.last_activity_at else "",
        "pools": pools,
        "stage_history": stage_history,
        "matches": match_rows,
        "tasks": [
            {
                "id": t.id,
                "task_type": t.task_type,
                "title": t.title,
                "due_at": t.due_at.isoformat() if t.due_at else "",
                "status": t.status,
            }
            for t in tasks
        ],
        "conversations": _build_conversations_payload(db, c),
    }


def _build_conversations_payload(db: Session, c: Candidate) -> list:
    """Order conversations by latest-message DESC, and messages inside each DESC (latest first)."""
    convs = list(c.conversations)

    def _conv_last_ts(conv):
        best = None
        for m in conv.messages:
            ts = _parse_msg_ts(m.timestamp_raw) or m.created_at
            if ts and (best is None or ts > best):
                best = ts
        return best or conv.captured_at or datetime.min

    convs.sort(key=_conv_last_ts, reverse=True)
    out = []
    for conv in convs:
        msgs_sorted = sorted(
            conv.messages,
            key=lambda m: _parse_msg_ts(m.timestamp_raw) or m.created_at or datetime.min,
            reverse=True,  # latest first
        )
        out.append({
            "id": conv.id,
            "captured_at": conv.captured_at.isoformat() if conv.captured_at else "",
            "summary": conv.summary,
            "sentiment": conv.sentiment,
            "person_brief": getattr(conv, "person_brief", None) or "",
            "suggested_pool_id": getattr(conv, "suggested_pool_id", None),
            "suggested_role_id": getattr(conv, "suggested_role_id", None),
            "pool_id": getattr(conv, "pool_id", None),
            "role_id": getattr(conv, "role_id", None),
            "analyzed_at": conv.analyzed_at.isoformat() if getattr(conv, "analyzed_at", None) else "",
            "analysis_status": getattr(conv, "analysis_status", None) or "pending",
            "analysis_error": getattr(conv, "analysis_error", None) or "",
            "message_count": len(msgs_sorted),
            "messages": [
                {
                    "sender": m.sender,
                    "body": m.body,
                    "timestamp": m.timestamp_raw,
                    "from_me": (m.sender or "").strip().lower() == "you",
                }
                for m in msgs_sorted
            ],
            "drafts": [
                {
                    "id": d.id,
                    "draft_text": d.draft_text,
                    "tone": d.tone,
                    "approved": d.approved,
                    "created_at": d.created_at.isoformat(),
                }
                for d in conv.drafts
            ],
            "classifications": [
                {
                    "id": rc.id,
                    "label": rc.label,
                    "confidence": rc.confidence,
                    "suggested_action": rc.suggested_action,
                    "suggested_reply": rc.suggested_reply,
                    "handled": rc.handled,
                    "created_at": rc.created_at.isoformat() if rc.created_at else "",
                }
                for rc in db.query(ReplyClassification)
                    .filter(ReplyClassification.conversation_id == conv.id)
                    .order_by(ReplyClassification.created_at.desc())
                    .all()
            ],
        })
    return out


@app.patch("/api/candidates/{candidate_id}")
def update_candidate(
    candidate_id: int,
    payload: NotesUpdate,
    db: Session = Depends(get_db),
    user: Optional[User] = Depends(current_user),
    _: bool = Depends(require_api_key),
):
    c = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not c:
        raise HTTPException(404, "Candidate not found")

    simple_map = {
        "notes": "notes", "full_name": "full_name", "headline": "headline",
        "current_title": "current_title", "current_company": "current_company",
        "location": "location", "about": "about", "email": "email", "phone": "phone",
        "visa_status": "visa_status", "salary_expectation": "salary_expectation",
        "owner_id": "owner_id", "years_experience": "years_experience",
    }
    for pyname, dbname in simple_map.items():
        v = getattr(payload, pyname)
        if v is not None:
            setattr(c, dbname, v)

    if payload.open_to_relocation is not None:
        c.open_to_relocation = bool(payload.open_to_relocation)

    if payload.skills is not None:
        c.skills_json = _jdump(payload.skills)
    if payload.languages is not None:
        c.languages_json = _jdump(payload.languages)

    if payload.tags is not None:
        c.tags = ", ".join(payload.tags) if payload.tags else ""

    if payload.stage is not None and payload.stage != (c.stage or ""):
        _record_stage_change(db, c.id, c.stage or "", payload.stage, user)
        c.stage = payload.stage

    if payload.pool_ids is not None:
        # Reset pool memberships
        c.pools = db.query(TalentPool).filter(TalentPool.id.in_(payload.pool_ids)).all()

    _touch_activity(c)
    db.commit()
    return {"ok": True}


# ========== Endpoints: manual candidate creation ==========

class CandidateCreateIn(BaseModel):
    full_name: str
    profile_url: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""
    headline: Optional[str] = ""
    current_title: Optional[str] = ""
    current_company: Optional[str] = ""
    location: Optional[str] = ""
    about: Optional[str] = ""
    skills: Optional[list] = None
    languages: Optional[list] = None
    visa_status: Optional[str] = ""
    open_to_relocation: Optional[bool] = False
    years_experience: Optional[int] = None
    salary_expectation: Optional[str] = ""
    tags: Optional[list] = None
    stage: Optional[str] = "New"
    notes: Optional[str] = ""
    source: Optional[str] = "manual"
    pool_ids: Optional[List[int]] = None
    conversation_transcript: Optional[str] = ""  # optional: one line per message "Sender: text"


@app.post("/api/candidates/create")
def create_candidate(
    payload: CandidateCreateIn,
    db: Session = Depends(get_db),
    user: Optional[User] = Depends(current_user),
    _: bool = Depends(require_api_key),
):
    """Create or upsert a candidate from a manual form. If conversation_transcript is given,
    each line 'Sender: body' becomes a message, and Claude classifies the latest candidate reply."""
    pu = (payload.profile_url or "").strip() or f"manual:{payload.full_name}:{int(datetime.utcnow().timestamp())}"

    # Dedup: if exact profile_url or email already exists, update instead.
    cand = upsert_candidate(
        db,
        profile_url=pu,
        full_name=payload.full_name,
        headline=payload.headline,
        current_title=payload.current_title,
        current_company=payload.current_company,
        location=payload.location,
        about=payload.about,
        email=payload.email,
        phone=payload.phone,
        skills=payload.skills,
        languages=payload.languages,
        visa_status=payload.visa_status,
        open_to_relocation=payload.open_to_relocation,
        years_experience=payload.years_experience,
        salary_expectation=payload.salary_expectation,
        notes=payload.notes,
        source=payload.source or "manual",
    )
    if payload.tags is not None:
        cand.tags = ", ".join(payload.tags) if payload.tags else ""
    if payload.stage:
        if (cand.stage or "") != payload.stage:
            _record_stage_change(db, cand.id, cand.stage or "", payload.stage, user)
        cand.stage = payload.stage
    if user and not cand.owner_id:
        cand.owner_id = user.id
    if payload.pool_ids is not None:
        cand.pools = db.query(TalentPool).filter(TalentPool.id.in_(payload.pool_ids)).all()

    db.commit()
    db.refresh(cand)

    conv_id = None
    if (payload.conversation_transcript or "").strip():
        conv = Conversation(
            candidate_id=cand.id,
            thread_url="",
            source="manual",
        )
        db.add(conv)
        db.flush()
        for ln in (payload.conversation_transcript or "").splitlines():
            if ":" not in ln:
                continue
            sender, body = ln.split(":", 1)
            db.add(Message(conversation_id=conv.id, sender=sender.strip(), body=body.strip()))
        db.commit()
        conv_id = conv.id
        # Auto-classify latest candidate reply (non-"You")
        try:
            msgs = db.query(Message).filter(Message.conversation_id == conv_id)\
                     .order_by(Message.id.asc()).all()
            last_cand_msg = next((m for m in reversed(msgs) if (m.sender or "").strip().lower() != "you"), None)
            if last_cand_msg:
                cls = ai_service.classify_reply(last_cand_msg.body, prior_context="")
                db.add(ReplyClassification(
                    conversation_id=conv_id,
                    last_message_body=last_cand_msg.body,
                    label=cls.get("label", "other"),
                    confidence=int(cls.get("confidence", 0) or 0),
                    suggested_action=cls.get("suggested_action", ""),
                    suggested_reply=cls.get("suggested_reply", ""),
                ))
                db.commit()
        except Exception as e:
            print(f"[warn] classify on manual create failed: {e}")

    return {"id": cand.id, "conversation_id": conv_id, "status": "created"}


# ========== Endpoints: CSV import ==========

@app.post("/api/candidates/import_csv")
async def import_candidates_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: Optional[User] = Depends(current_user),
    _: bool = Depends(require_api_key),
):
    """Bulk import candidates from CSV. Recognized headers (case-insensitive):
    full_name, name, profile_url, linkedin_url, linkedin, email, phone, title, current_title,
    company, current_company, location, headline, about, skills, languages, visa_status,
    years_experience, salary_expectation, tags, stage, pool.
    """
    content = (await file.read()).decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(content))
    created = 0
    updated = 0
    failed = 0
    for row in reader:
        try:
            # Normalize keys to lowercase
            r = { (k or "").strip().lower(): (v or "").strip() for k, v in row.items() }
            name = r.get("full_name") or r.get("name") or ""
            if not name:
                failed += 1; continue
            pu = r.get("profile_url") or r.get("linkedin_url") or r.get("linkedin") or ""
            existed = False
            if pu:
                existed = db.query(Candidate).filter(Candidate.profile_url == pu).first() is not None
            elif r.get("email"):
                existed = db.query(Candidate).filter(Candidate.email == r["email"]).first() is not None

            cand = upsert_candidate(
                db,
                profile_url=pu or f"csv:{name}:{created+updated}",
                full_name=name,
                headline=r.get("headline", ""),
                current_title=r.get("current_title") or r.get("title", ""),
                current_company=r.get("current_company") or r.get("company", ""),
                location=r.get("location", ""),
                about=r.get("about", ""),
                email=r.get("email", ""),
                phone=r.get("phone", ""),
                visa_status=r.get("visa_status", ""),
                salary_expectation=r.get("salary_expectation", ""),
                skills=_coerce_list(r.get("skills", "")),
                languages=_coerce_list(r.get("languages", "")),
                years_experience=int(r["years_experience"]) if r.get("years_experience", "").isdigit() else None,
                source="csv",
            )
            if r.get("stage"):
                if (cand.stage or "") != r["stage"]:
                    _record_stage_change(db, cand.id, cand.stage or "", r["stage"], user)
                cand.stage = r["stage"]
            if r.get("tags"):
                cand.tags = r["tags"]
            if user and not cand.owner_id:
                cand.owner_id = user.id
            # Optional pool by name
            pool_name = r.get("pool") or r.get("talent_pool")
            if pool_name:
                pool = db.query(TalentPool).filter(TalentPool.name == pool_name).first()
                if not pool:
                    pool = TalentPool(name=pool_name, owner_id=user.id if user else None)
                    db.add(pool); db.flush()
                if pool not in cand.pools:
                    cand.pools.append(pool)
            db.commit()
            if existed: updated += 1
            else: created += 1
        except Exception as e:
            print(f"[warn] csv row failed: {e}")
            db.rollback()
            failed += 1
    return {"created": created, "updated": updated, "failed": failed}


# ========== Endpoints: Apify profile enrichment (replaces Proxycurl) ==========

class EnrichIn(BaseModel):
    linkedin_url: Optional[str] = ""
    candidate_id: Optional[int] = None
    save: Optional[bool] = True  # when False, runs the actor and returns the profile but does not upsert


@app.post("/api/candidates/enrich")
def enrich_candidate(
    payload: EnrichIn,
    db: Session = Depends(get_db),
    user: Optional[User] = Depends(current_user),
    _: bool = Depends(require_api_key),
):
    """Enrich (or create) a candidate from a LinkedIn URL via an Apify actor.

    Default actor is harvestapi/linkedin-profile-scraper (works on free Apify plan).
    Users can override APIFY_ACTOR_ID in Settings; we send the union of common
    input shapes so multiple actors can be used.
    """
    key = (get_setting(db, "apify_api_key", "")
           or get_setting(db, "proxycurl_api_key", ""))
    if not key:
        raise HTTPException(400, "Apify API key is not set. Open Settings to add it.")
    actor_id = get_setting(db, "apify_actor_id", "") or ai_service.APIFY_DEFAULT_ACTOR

    cand = None
    url = (payload.linkedin_url or "").strip()
    if payload.candidate_id:
        cand = db.query(Candidate).filter(Candidate.id == payload.candidate_id).first()
        if cand and not url:
            url = cand.profile_url or ""
    if not url:
        raise HTTPException(400, "No LinkedIn URL provided.")

    try:
        data = ai_service.apify_enrich_profile(url, key, actor_id=actor_id)
    except Exception as e:
        raise HTTPException(502, f"Apify error: {e}")

    profile_out = {k: v for k, v in data.items() if k != "_raw"}

    if not payload.save:
        return {"profile": profile_out, "saved": False}

    cand = upsert_candidate(
        db,
        profile_url=data.get("profile_url") or url,
        full_name=data.get("full_name") or "Unknown",
        headline=data.get("headline"),
        current_title=data.get("current_title"),
        current_company=data.get("current_company"),
        location=data.get("location"),
        about=data.get("about"),
        skills=data.get("skills"),
        languages=data.get("languages"),
        email=data.get("email"),
        phone=data.get("phone"),
        source="apify",
    )
    if user and not cand.owner_id:
        cand.owner_id = user.id
    db.commit()
    return {"id": cand.id, "full_name": cand.full_name, "source": "apify", "profile": profile_out, "saved": True}


# ========== Endpoints: AI candidate brief ==========

@app.post("/api/candidates/{candidate_id}/brief")
def candidate_brief(
    candidate_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    c = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not c:
        raise HTTPException(404, "Candidate not found")
    conversations = [
        {
            "captured_at": conv.captured_at.isoformat() if conv.captured_at else "",
            "summary": conv.summary or "",
            "sentiment": conv.sentiment or "",
        }
        for conv in c.conversations
    ]
    matches = []
    for m in db.query(CandidateRoleMatch).filter(CandidateRoleMatch.candidate_id == c.id).all():
        r = m.role
        matches.append({
            "role_title": r.title if r else "",
            "role_company": (r.company if r else "") or "",
            "fit_score": m.fit_score,
            "fit_reason": m.fit_reason or "",
        })
    try:
        brief = ai_service.write_candidate_brief(_candidate_summary(c), conversations, matches)
    except Exception as e:
        raise HTTPException(500, f"Brief generation failed: {e}")
    return brief


# ========== Endpoints: Talent Pools ==========

class PoolIn(BaseModel):
    name: str
    description: Optional[str] = ""
    color: Optional[str] = "#2563eb"
    role_id: Optional[int] = None


@app.get("/api/pools")
def list_pools(
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    rows = db.query(TalentPool).order_by(TalentPool.updated_at.desc()).all()
    return [
        {
            "id": p.id, "name": p.name, "description": p.description or "",
            "color": p.color or "#2563eb", "role_id": p.role_id,
            "owner_id": p.owner_id,
            "count": len(p.candidates),
            "created_at": p.created_at.isoformat() if p.created_at else "",
        }
        for p in rows
    ]


@app.post("/api/pools")
def create_pool(
    payload: PoolIn,
    db: Session = Depends(get_db),
    user: Optional[User] = Depends(current_user),
    _: bool = Depends(require_api_key),
):
    p = TalentPool(
        name=payload.name,
        description=payload.description or "",
        color=payload.color or "#2563eb",
        role_id=payload.role_id,
        owner_id=user.id if user else None,
    )
    db.add(p); db.commit(); db.refresh(p)
    return {"id": p.id, "name": p.name}


@app.patch("/api/pools/{pool_id}")
def update_pool(
    pool_id: int, payload: PoolIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    p = db.query(TalentPool).filter(TalentPool.id == pool_id).first()
    if not p:
        raise HTTPException(404, "Pool not found")
    p.name = payload.name
    p.description = payload.description or ""
    p.color = payload.color or "#2563eb"
    p.role_id = payload.role_id
    db.commit()
    return {"ok": True}


@app.delete("/api/pools/{pool_id}")
def delete_pool(
    pool_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    p = db.query(TalentPool).filter(TalentPool.id == pool_id).first()
    if not p:
        raise HTTPException(404, "Pool not found")
    db.delete(p); db.commit()
    return {"ok": True}


class PoolAssignIn(BaseModel):
    candidate_ids: List[int]


@app.post("/api/pools/{pool_id}/add")
def add_to_pool(
    pool_id: int, payload: PoolAssignIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    p = db.query(TalentPool).filter(TalentPool.id == pool_id).first()
    if not p:
        raise HTTPException(404, "Pool not found")
    added = 0
    for cid in payload.candidate_ids:
        c = db.query(Candidate).filter(Candidate.id == cid).first()
        if c and p not in c.pools:
            c.pools.append(p)
            added += 1
    db.commit()
    return {"added": added}


@app.post("/api/pools/{pool_id}/remove")
def remove_from_pool(
    pool_id: int, payload: PoolAssignIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    p = db.query(TalentPool).filter(TalentPool.id == pool_id).first()
    if not p:
        raise HTTPException(404, "Pool not found")
    removed = 0
    for cid in payload.candidate_ids:
        c = db.query(Candidate).filter(Candidate.id == cid).first()
        if c and p in c.pools:
            c.pools.remove(p)
            removed += 1
    db.commit()
    return {"removed": removed}


# Autocomplete endpoints for tags, skills, companies, locations
@app.get("/api/autocomplete/{field}")
def autocomplete(
    field: str,
    q: str = "",
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    field = field.lower()
    values = set()
    if field == "tags":
        for c in db.query(Candidate.tags).all():
            for t in (c[0] or "").split(","):
                t = t.strip()
                if t and (not q or q.lower() in t.lower()):
                    values.add(t)
    elif field == "skills":
        for c in db.query(Candidate.skills_json).all():
            for s in _jload(c[0], []):
                if s and (not q or q.lower() in s.lower()):
                    values.add(s)
    elif field == "companies":
        for c in db.query(Candidate.current_company).distinct().all():
            if c[0] and (not q or q.lower() in c[0].lower()):
                values.add(c[0])
    elif field == "locations":
        for c in db.query(Candidate.location).distinct().all():
            if c[0] and (not q or q.lower() in c[0].lower()):
                values.add(c[0])
    else:
        raise HTTPException(404, "Unknown autocomplete field")
    return sorted(values)[:50]


# ========== Endpoints: AI drafting ==========

@app.post("/api/conversations/{conversation_id}/draft")
def create_draft(
    conversation_id: int,
    payload: DraftRequest,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    conv = db.query(Conversation).filter(Conversation.id == conversation_id).first()
    if not conv:
        raise HTTPException(404, "Conversation not found")

    try:
        draft_text = ai_service.draft_reply(
            [{"sender": m.sender, "body": m.body} for m in conv.messages],
            conv.candidate.full_name,
            tone=payload.tone,
            recruiter_goal=payload.goal,
        )
    except Exception as e:
        raise HTTPException(500, f"AI draft failed: {e}")

    draft = ReplyDraft(
        conversation_id=conv.id,
        draft_text=draft_text,
        tone=payload.tone,
    )
    db.add(draft)
    db.commit()
    return {
        "draft_id": draft.id,
        "draft_text": draft_text,
        "tone": payload.tone,
    }


@app.post("/api/drafts/{draft_id}/approve")
def approve_draft(
    draft_id: int,
    payload: DraftApproval,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    d = db.query(ReplyDraft).filter(ReplyDraft.id == draft_id).first()
    if not d:
        raise HTTPException(404, "Draft not found")
    d.approved = payload.approved
    d.approved_at = datetime.utcnow() if payload.approved else None
    db.commit()
    return {"ok": True, "approved": d.approved}


# ========== Endpoints: Roles (Role Intake) ==========

class RoleCreateIn(BaseModel):
    jd_text: str = ""
    title: Optional[str] = ""
    company: Optional[str] = ""
    location: Optional[str] = ""
    seniority: Optional[str] = ""
    employment_type: Optional[str] = ""
    hiring_notes: Optional[str] = ""


class RolePatchIn(BaseModel):
    title: Optional[str] = None
    company: Optional[str] = None
    location: Optional[str] = None
    seniority: Optional[str] = None
    employment_type: Optional[str] = None
    jd_text: Optional[str] = None
    must_have: Optional[list] = None
    nice_to_have: Optional[list] = None
    target_companies: Optional[list] = None
    target_titles: Optional[list] = None
    search_keywords: Optional[list] = None
    outreach_angle: Optional[str] = None
    persona: Optional[str] = None
    exclusions: Optional[str] = None
    status: Optional[str] = None


@app.post("/api/roles")
def create_role(
    payload: RoleCreateIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    """Create a role. If jd_text is given, Claude parses it into structured fields."""
    parsed = {}
    if payload.jd_text.strip():
        try:
            parsed = ai_service.parse_jd(payload.jd_text, payload.hiring_notes or "")
        except Exception as e:
            print(f"[warn] JD parse failed: {e}")

    role = Role(
        title=payload.title or parsed.get("title") or "Untitled role",
        company=payload.company or parsed.get("company") or "",
        location=payload.location or parsed.get("location") or "",
        seniority=payload.seniority or parsed.get("seniority") or "",
        employment_type=payload.employment_type or parsed.get("employment_type") or "",
        jd_text=payload.jd_text or "",
        must_have_json=_jdump(parsed.get("must_have", [])),
        nice_to_have_json=_jdump(parsed.get("nice_to_have", [])),
        target_companies_json=_jdump(parsed.get("target_companies", [])),
        target_titles_json=_jdump(parsed.get("target_titles", [])),
        search_keywords_json=_jdump(parsed.get("search_keywords", [])),
        outreach_angle=parsed.get("outreach_angle", ""),
        persona=parsed.get("persona", ""),
        exclusions=_jdump(parsed.get("exclusions", [])) if isinstance(parsed.get("exclusions"), list) else (parsed.get("exclusions") or ""),
        status="open",
    )
    db.add(role)
    db.commit()
    db.refresh(role)

    # Seed a default template library for this role.
    seed_defaults = [
        ("Connection note", "connection_note",
         "Hi {{first_name}}, I came across your profile — your work at {{current_company}} looks strongly aligned with a {{role_title}} role I'm filling. Open to a short chat?"),
        ("First outreach", "first_outreach",
         "Hi {{first_name}},\n\nI'm Habib at TMC — I'm currently searching for a {{role_title}} and your experience as {{current_title}} at {{current_company}} caught my attention, specifically {{matched_skill}}.\n\nThe role is {{role_location}}, working on {{outreach_angle}}.\n\nWould you be open to a 15-minute call this week to see if it's a fit?\n\nBest,\nHabib"),
        ("Follow-up #1", "follow_up_1",
         "Hi {{first_name}}, bumping this in case it got lost. Still keen to chat about the {{role_title}} role if it's of interest. No pressure either way."),
        ("Follow-up #2", "follow_up_2",
         "Hi {{first_name}}, last message from me on this one. If the timing's wrong, totally understand — happy to stay in touch for future roles."),
        ("Close the loop", "close_loop",
         "Thanks {{first_name}} — appreciate you taking a look. I'll close the loop here, but if things change on your side, my door's always open."),
    ]
    for name, ttype, body in seed_defaults:
        db.add(MessageTemplate(
            role_id=role.id,
            name=name,
            template_type=ttype,
            channel="linkedin",
            body_template=body,
        ))
    db.commit()
    db.refresh(role)
    return _role_dict(role)


@app.get("/api/roles")
def list_roles(
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    rows = db.query(Role).order_by(Role.updated_at.desc()).all()
    return [_role_dict(r) for r in rows]


@app.get("/api/roles/{role_id}")
def get_role(
    role_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    r = db.query(Role).filter(Role.id == role_id).first()
    if not r:
        raise HTTPException(404, "Role not found")

    # Attach candidate matches (sorted by score desc) + templates.
    matches = db.query(CandidateRoleMatch).filter(
        CandidateRoleMatch.role_id == role_id
    ).all()
    match_rows = []
    for m in matches:
        c = m.candidate
        match_rows.append({
            "id": m.id,
            "candidate_id": m.candidate_id,
            "candidate_name": c.full_name if c else "",
            "current_title": c.current_title if c else "",
            "current_company": c.current_company if c else "",
            "fit_score": m.fit_score,
            "fit_reason": m.fit_reason or "",
            "fit_bullets": _jload(m.fit_bullets_json, []),
            "risk_flags": _jload(m.risk_flags_json, []),
            "stage": m.stage or "New",
        })
    match_rows.sort(key=lambda x: (x["fit_score"] or 0), reverse=True)

    templates = db.query(MessageTemplate).filter(
        MessageTemplate.role_id == role_id
    ).order_by(MessageTemplate.id.asc()).all()

    d = _role_dict(r)
    d["matches"] = match_rows
    d["templates"] = [
        {
            "id": t.id,
            "name": t.name,
            "template_type": t.template_type,
            "channel": t.channel,
            "body_template": t.body_template,
            "subject_template": t.subject_template or "",
        }
        for t in templates
    ]
    return d


@app.patch("/api/roles/{role_id}")
def patch_role(
    role_id: int,
    payload: RolePatchIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    r = db.query(Role).filter(Role.id == role_id).first()
    if not r:
        raise HTTPException(404, "Role not found")

    simple_fields = [
        "title", "company", "location", "seniority", "employment_type",
        "jd_text", "outreach_angle", "persona", "status"
    ]
    for f in simple_fields:
        v = getattr(payload, f)
        if v is not None:
            setattr(r, f, v)

    json_fields = {
        "must_have": "must_have_json",
        "nice_to_have": "nice_to_have_json",
        "target_companies": "target_companies_json",
        "target_titles": "target_titles_json",
        "search_keywords": "search_keywords_json",
    }
    for pyname, dbname in json_fields.items():
        v = getattr(payload, pyname)
        if v is not None:
            setattr(r, dbname, _jdump(v))

    if payload.exclusions is not None:
        r.exclusions = payload.exclusions

    db.commit()
    return _role_dict(r)


@app.delete("/api/roles/{role_id}")
def delete_role(
    role_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    r = db.query(Role).filter(Role.id == role_id).first()
    if not r:
        raise HTTPException(404, "Role not found")
    db.delete(r)
    db.commit()
    return {"ok": True}


@app.post("/api/roles/{role_id}/reparse")
def reparse_role(
    role_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    """Re-run JD parser on the saved jd_text (useful after editing)."""
    r = db.query(Role).filter(Role.id == role_id).first()
    if not r:
        raise HTTPException(404, "Role not found")
    if not (r.jd_text or "").strip():
        raise HTTPException(400, "Role has no jd_text to parse")
    parsed = ai_service.parse_jd(r.jd_text, "")
    r.must_have_json = _jdump(parsed.get("must_have", []))
    r.nice_to_have_json = _jdump(parsed.get("nice_to_have", []))
    r.target_companies_json = _jdump(parsed.get("target_companies", []))
    r.target_titles_json = _jdump(parsed.get("target_titles", []))
    r.search_keywords_json = _jdump(parsed.get("search_keywords", []))
    r.outreach_angle = parsed.get("outreach_angle", r.outreach_angle or "")
    r.persona = parsed.get("persona", r.persona or "")
    db.commit()
    return _role_dict(r)


# ========== Endpoints: Fit scoring ==========

class ScoreIn(BaseModel):
    candidate_id: int


@app.post("/api/roles/{role_id}/score")
def score_candidate_for_role(
    role_id: int,
    payload: ScoreIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    r = db.query(Role).filter(Role.id == role_id).first()
    if not r:
        raise HTTPException(404, "Role not found")
    c = db.query(Candidate).filter(Candidate.id == payload.candidate_id).first()
    if not c:
        raise HTTPException(404, "Candidate not found")

    role_dict = _role_dict(r)
    cand_dict = _candidate_summary(c)

    try:
        result = ai_service.score_fit(role_dict, cand_dict)
    except Exception as e:
        raise HTTPException(500, f"Fit scoring failed: {e}")

    match = db.query(CandidateRoleMatch).filter(
        CandidateRoleMatch.role_id == role_id,
        CandidateRoleMatch.candidate_id == c.id,
    ).first()
    if not match:
        match = CandidateRoleMatch(role_id=role_id, candidate_id=c.id)
        db.add(match)

    match.fit_score = int(result.get("score", 0) or 0)
    match.fit_reason = result.get("reason", "")
    match.fit_bullets_json = _jdump(result.get("bullets", []))
    match.risk_flags_json = _jdump(result.get("risks", []))
    db.commit()
    db.refresh(match)
    return {
        "match_id": match.id,
        "fit_score": match.fit_score,
        "fit_reason": match.fit_reason,
        "fit_bullets": _jload(match.fit_bullets_json, []),
        "risk_flags": _jload(match.risk_flags_json, []),
    }


@app.post("/api/roles/{role_id}/score_all")
def score_all_candidates_for_role(
    role_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    """Score every candidate against this role. Slow — runs one AI call per candidate."""
    r = db.query(Role).filter(Role.id == role_id).first()
    if not r:
        raise HTTPException(404, "Role not found")

    role_dict = _role_dict(r)
    candidates = db.query(Candidate).all()
    scored = 0
    failed = 0
    for c in candidates:
        try:
            result = ai_service.score_fit(role_dict, _candidate_summary(c))
            match = db.query(CandidateRoleMatch).filter(
                CandidateRoleMatch.role_id == role_id,
                CandidateRoleMatch.candidate_id == c.id,
            ).first()
            if not match:
                match = CandidateRoleMatch(role_id=role_id, candidate_id=c.id)
                db.add(match)
            match.fit_score = int(result.get("score", 0) or 0)
            match.fit_reason = result.get("reason", "")
            match.fit_bullets_json = _jdump(result.get("bullets", []))
            match.risk_flags_json = _jdump(result.get("risks", []))
            db.commit()
            scored += 1
        except Exception as e:
            print(f"[warn] score failed for candidate {c.id}: {e}")
            failed += 1
    return {"scored": scored, "failed": failed, "total_candidates": len(candidates)}


class MatchStagePatch(BaseModel):
    stage: Optional[str] = None


@app.patch("/api/matches/{match_id}")
def patch_match(
    match_id: int,
    payload: MatchStagePatch,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    m = db.query(CandidateRoleMatch).filter(CandidateRoleMatch.id == match_id).first()
    if not m:
        raise HTTPException(404, "Match not found")
    if payload.stage is not None:
        m.stage = payload.stage
    db.commit()
    return {"ok": True, "stage": m.stage}


# ========== Endpoints: Templates ==========

class TemplateIn(BaseModel):
    role_id: Optional[int] = None
    name: str
    template_type: str = "first_outreach"
    channel: str = "linkedin"
    subject_template: Optional[str] = ""
    body_template: str


@app.get("/api/templates")
def list_templates(
    role_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    q = db.query(MessageTemplate)
    if role_id is not None:
        q = q.filter(MessageTemplate.role_id == role_id)
    rows = q.order_by(MessageTemplate.id.asc()).all()
    return [
        {
            "id": t.id,
            "role_id": t.role_id,
            "name": t.name,
            "template_type": t.template_type,
            "channel": t.channel,
            "subject_template": t.subject_template or "",
            "body_template": t.body_template,
        }
        for t in rows
    ]


@app.post("/api/templates")
def create_template(
    payload: TemplateIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    t = MessageTemplate(
        role_id=payload.role_id,
        name=payload.name,
        template_type=payload.template_type,
        channel=payload.channel,
        subject_template=payload.subject_template or "",
        body_template=payload.body_template,
    )
    db.add(t)
    db.commit()
    db.refresh(t)
    return {"id": t.id}


@app.patch("/api/templates/{template_id}")
def update_template(
    template_id: int,
    payload: TemplateIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    t = db.query(MessageTemplate).filter(MessageTemplate.id == template_id).first()
    if not t:
        raise HTTPException(404, "Template not found")
    t.role_id = payload.role_id
    t.name = payload.name
    t.template_type = payload.template_type
    t.channel = payload.channel
    t.subject_template = payload.subject_template or ""
    t.body_template = payload.body_template
    db.commit()
    return {"ok": True}


@app.delete("/api/templates/{template_id}")
def delete_template(
    template_id: int,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    t = db.query(MessageTemplate).filter(MessageTemplate.id == template_id).first()
    if not t:
        raise HTTPException(404, "Template not found")
    db.delete(t)
    db.commit()
    return {"ok": True}


class TemplateDraftIn(BaseModel):
    template_id: int
    candidate_id: int
    tone: str = "professional"


@app.post("/api/templates/draft")
def draft_from_template(
    payload: TemplateDraftIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    """Render a template with Claude, substituting variables and tightening phrasing."""
    t = db.query(MessageTemplate).filter(MessageTemplate.id == payload.template_id).first()
    if not t:
        raise HTTPException(404, "Template not found")
    c = db.query(Candidate).filter(Candidate.id == payload.candidate_id).first()
    if not c:
        raise HTTPException(404, "Candidate not found")

    role = t.role
    role_dict = _role_dict(role) if role else {"title": "", "company": "", "must_have": []}
    first_name = (c.full_name or "").split(" ")[0] or "there"
    variables = {
        "first_name": first_name,
        "full_name": c.full_name or "",
        "current_company": c.current_company or "",
        "current_title": c.current_title or "",
        "role_title": role_dict.get("title", ""),
        "role_company": role_dict.get("company", ""),
        "role_location": role_dict.get("location", ""),
        "outreach_angle": role_dict.get("outreach_angle", ""),
        "matched_skill": (role_dict.get("must_have", []) or [""])[0],
    }
    pre_filled = ai_service.render_template(t.body_template, variables)

    try:
        final = ai_service.personalize_outreach(
            role=role_dict,
            candidate=_candidate_summary(c),
            template_body=pre_filled,
            template_type=t.template_type or "first_outreach",
            tone=payload.tone,
        )
    except Exception as e:
        raise HTTPException(500, f"Template draft failed: {e}")

    return {
        "draft_text": final,
        "variables_used": variables,
        "pre_filled": pre_filled,
    }


# ========== Endpoints: Reply Triage ==========

@app.get("/api/replies/triage")
def list_reply_triage(
    handled: bool = False,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    """All reply classifications that still need attention (or all if handled=true)."""
    q = db.query(ReplyClassification)
    if not handled:
        q = q.filter(ReplyClassification.handled.is_(False))
    rows = q.order_by(ReplyClassification.created_at.desc()).all()
    out = []
    for rc in rows:
        conv = rc.conversation
        cand = conv.candidate if conv else None
        out.append({
            "id": rc.id,
            "conversation_id": rc.conversation_id,
            "candidate_id": cand.id if cand else None,
            "candidate_name": cand.full_name if cand else "",
            "current_title": cand.current_title if cand else "",
            "current_company": cand.current_company if cand else "",
            "label": rc.label,
            "confidence": rc.confidence,
            "suggested_action": rc.suggested_action or "",
            "suggested_reply": rc.suggested_reply or "",
            "last_message_body": rc.last_message_body or "",
            "handled": rc.handled,
            "created_at": rc.created_at.isoformat() if rc.created_at else "",
        })
    return out


class HandledIn(BaseModel):
    handled: bool = True


@app.patch("/api/replies/triage/{rc_id}")
def patch_reply_triage(
    rc_id: int,
    payload: HandledIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    rc = db.query(ReplyClassification).filter(ReplyClassification.id == rc_id).first()
    if not rc:
        raise HTTPException(404, "Reply classification not found")
    rc.handled = bool(payload.handled)
    db.commit()
    return {"ok": True}


# ========== Endpoints: Tasks / Follow-ups ==========

class TaskIn(BaseModel):
    entity_type: str
    entity_id: int
    task_type: str = "follow_up"
    title: str
    due_at: Optional[str] = None  # ISO string


class TaskPatchIn(BaseModel):
    status: Optional[str] = None
    due_at: Optional[str] = None
    title: Optional[str] = None


@app.get("/api/tasks")
def list_tasks(
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    q = db.query(Task)
    if status:
        q = q.filter(Task.status == status)
    rows = q.order_by(Task.due_at.asc()).all()
    out = []
    now = datetime.utcnow()
    for t in rows:
        linked_name = ""
        linked_profile = ""
        if t.entity_type == "candidate":
            c = db.query(Candidate).filter(Candidate.id == t.entity_id).first()
            if c:
                linked_name = c.full_name
                linked_profile = c.profile_url or ""
        out.append({
            "id": t.id,
            "entity_type": t.entity_type,
            "entity_id": t.entity_id,
            "task_type": t.task_type,
            "title": t.title,
            "due_at": t.due_at.isoformat() if t.due_at else "",
            "status": t.status,
            "overdue": bool(t.due_at and t.status == "pending" and t.due_at < now),
            "candidate_name": linked_name,
            "candidate_profile_url": linked_profile,
        })
    return out


@app.post("/api/tasks")
def create_task(
    payload: TaskIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    due = None
    if payload.due_at:
        try:
            due = datetime.fromisoformat(payload.due_at.replace("Z", ""))
        except Exception:
            due = None
    t = Task(
        entity_type=payload.entity_type,
        entity_id=payload.entity_id,
        task_type=payload.task_type,
        title=payload.title,
        due_at=due,
    )
    db.add(t)
    db.commit()
    db.refresh(t)
    return {"id": t.id}


@app.patch("/api/tasks/{task_id}")
def patch_task(
    task_id: int,
    payload: TaskPatchIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    t = db.query(Task).filter(Task.id == task_id).first()
    if not t:
        raise HTTPException(404, "Task not found")
    if payload.status is not None:
        t.status = payload.status
    if payload.due_at is not None:
        try:
            t.due_at = datetime.fromisoformat(payload.due_at.replace("Z", ""))
        except Exception:
            pass
    if payload.title is not None:
        t.title = payload.title
    db.commit()
    return {"ok": True}


# ========== Endpoints: Auth (login / logout / current user) ==========

class LoginIn(BaseModel):
    username: str
    password: str


@app.post("/api/auth/login")
def auth_login(payload: LoginIn, db: Session = Depends(get_db)):
    u = db.query(User).filter(User.username == payload.username).first()
    if not u or not ai_service.verify_password(payload.password, u.password_hash):
        raise HTTPException(401, "Invalid username or password")
    if not u.is_active:
        raise HTTPException(403, "Your account is pending admin approval")
    if not u.api_key:
        u.api_key = _new_user_api_key()
    token = ai_service.new_session_token()
    db.add(SessionRow(
        token=token, user_id=u.id,
        expires_at=datetime.utcnow() + timedelta(days=30),
    ))
    db.commit()
    return {"token": token, "user": _user_dict(u, reveal_key=True)}


@app.post("/api/auth/logout")
def auth_logout(
    x_session_token: str = Header(None, alias="X-Session-Token"),
    db: Session = Depends(get_db),
):
    if x_session_token:
        s = db.query(SessionRow).filter(SessionRow.token == x_session_token).first()
        if s:
            db.delete(s); db.commit()
    return {"ok": True}


@app.get("/api/auth/me")
def auth_me(
    user: Optional[User] = Depends(current_user),
    db: Session = Depends(get_db),
):
    users = [_user_dict(u) for u in db.query(User).filter(User.is_active.is_(True)).all()]
    # Reveal the logged-in user's own api_key so the dashboard can show it.
    return {"user": _user_dict(user, reveal_key=True), "all_users": users}


class RegisterIn(BaseModel):
    username: str
    password: str
    display_name: Optional[str] = ""
    email: Optional[str] = ""
    role: Optional[str] = "recruiter"


@app.post("/api/auth/users")
def create_user(
    payload: RegisterIn,
    db: Session = Depends(get_db),
    user: Optional[User] = Depends(current_user),
    _: bool = Depends(require_api_key),
):
    """Create a teammate. Only admins can create users (falls back to first-run seed)."""
    if user and user.role != "admin" and not getattr(user, "is_admin", False):
        raise HTTPException(403, "Only admins can add users")
    if db.query(User).filter(User.username == payload.username).first():
        raise HTTPException(400, "Username already exists")
    u = User(
        username=payload.username,
        password_hash=ai_service.hash_password(payload.password),
        display_name=payload.display_name or payload.username,
        email=payload.email or "",
        role=payload.role or "recruiter",
        is_active=True,
        is_admin=(payload.role == "admin"),
        api_key=_new_user_api_key(),
    )
    db.add(u); db.commit(); db.refresh(u)
    return {"id": u.id, "username": u.username, "api_key": u.api_key}


# ---------- Public self-registration (toggleable) ----------

class PublicRegisterIn(BaseModel):
    username: str
    password: str
    display_name: Optional[str] = ""
    email: Optional[str] = ""


@app.post("/api/auth/register")
def public_register(payload: PublicRegisterIn, db: Session = Depends(get_db)):
    """Self-service signup. Disabled by default — an admin has to enable it
    via Settings → 'allow_self_register'. New users are created as pending
    (is_active=False) unless 'auto_approve_signups' is also on."""
    if get_setting(db, "allow_self_register", "0") != "1":
        raise HTTPException(403, "Self-registration is disabled on this workspace.")
    if not payload.username or not payload.password:
        raise HTTPException(400, "username and password required")
    if db.query(User).filter(User.username == payload.username).first():
        raise HTTPException(400, "Username already exists")
    auto_approve = get_setting(db, "auto_approve_signups", "0") == "1"
    u = User(
        username=payload.username,
        password_hash=ai_service.hash_password(payload.password),
        display_name=payload.display_name or payload.username,
        email=payload.email or "",
        role="recruiter",
        is_active=auto_approve,
        is_admin=False,
        api_key=_new_user_api_key(),
    )
    db.add(u); db.commit(); db.refresh(u)
    return {
        "id": u.id,
        "username": u.username,
        "is_active": u.is_active,
        "needs_approval": not auto_approve,
    }


# ---------- Per-user API key for the Chrome extension ----------

@app.get("/api/auth/api_key")
def get_my_api_key(
    user: Optional[User] = Depends(current_user),
    db: Session = Depends(get_db),
):
    """Returns the logged-in user's personal extension key.
    Only available when authenticated via session token (not the key itself)."""
    if not user:
        raise HTTPException(401, "Log in first")
    if not user.api_key:
        user.api_key = _new_user_api_key()
        db.commit()
    return {"api_key": user.api_key, "user_id": user.id, "username": user.username}


@app.post("/api/auth/api_key/rotate")
def rotate_my_api_key(
    user: Optional[User] = Depends(current_user),
    db: Session = Depends(get_db),
):
    """Invalidate the old key and mint a new one. Existing Chrome extensions
    stop working until the recruiter pastes the new key."""
    if not user:
        raise HTTPException(401, "Log in first")
    user.api_key = _new_user_api_key()
    db.commit()
    return {"api_key": user.api_key}


# ---------- Admin dashboard: list / toggle / promote users ----------

@app.get("/api/admin/users")
def admin_list_users(
    db: Session = Depends(get_db),
    _admin: Optional[User] = Depends(require_admin),
    __: bool = Depends(require_api_key),
):
    """Admin-only. Returns every user + their usage (candidates & conversations owned)."""
    users = db.query(User).order_by(User.created_at.desc()).all()
    out = []
    for u in users:
        cand_count = db.query(func.count(Candidate.id)).filter(Candidate.owner_user_id == u.id).scalar() or 0
        conv_count = db.query(func.count(Conversation.id)).filter(Conversation.owner_user_id == u.id).scalar() or 0
        d = _user_dict(u)
        d["candidate_count"] = int(cand_count)
        d["conversation_count"] = int(conv_count)
        out.append(d)
    return {"users": out, "total": len(out)}


class AdminUserPatchIn(BaseModel):
    is_active: Optional[bool] = None
    is_admin: Optional[bool] = None
    role: Optional[str] = None
    display_name: Optional[str] = None
    reset_password: Optional[str] = None


@app.patch("/api/admin/users/{user_id}")
def admin_update_user(
    user_id: int,
    payload: AdminUserPatchIn,
    db: Session = Depends(get_db),
    _admin: Optional[User] = Depends(require_admin),
    __: bool = Depends(require_api_key),
):
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(404, "User not found")
    if payload.is_active is not None:
        u.is_active = bool(payload.is_active)
    if payload.is_admin is not None:
        u.is_admin = bool(payload.is_admin)
        if payload.is_admin:
            u.role = "admin"
    if payload.role is not None:
        u.role = payload.role
    if payload.display_name is not None:
        u.display_name = payload.display_name
    if payload.reset_password:
        u.password_hash = ai_service.hash_password(payload.reset_password)
    db.commit()
    return _user_dict(u)


@app.get("/api/admin/users/{user_id}/api_key")
def admin_get_user_api_key(
    user_id: int,
    db: Session = Depends(get_db),
    _admin: Optional[User] = Depends(require_admin),
    __: bool = Depends(require_api_key),
):
    """Admin-only. Reveal a teammate's personal extension key so admin can
    hand it to them securely (or verify their extension is configured right)."""
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(404, "User not found")
    if not u.api_key:
        u.api_key = _new_user_api_key()
        db.commit()
    return {"user_id": u.id, "username": u.username, "api_key": u.api_key}


@app.post("/api/admin/users/{user_id}/api_key/rotate")
def admin_rotate_user_api_key(
    user_id: int,
    db: Session = Depends(get_db),
    _admin: Optional[User] = Depends(require_admin),
    __: bool = Depends(require_api_key),
):
    """Admin-only. Invalidate a teammate's old key and mint a new one.
    Their existing extension stops working until they paste the new key."""
    u = db.query(User).filter(User.id == user_id).first()
    if not u:
        raise HTTPException(404, "User not found")
    u.api_key = _new_user_api_key()
    db.commit()
    return {"user_id": u.id, "username": u.username, "api_key": u.api_key}


@app.get("/api/admin/stats")
def admin_stats(
    db: Session = Depends(get_db),
    _admin: Optional[User] = Depends(require_admin),
    __: bool = Depends(require_api_key),
):
    """High-level numbers the admin sees on their dashboard."""
    return {
        "users_total": db.query(func.count(User.id)).scalar() or 0,
        "users_active": db.query(func.count(User.id)).filter(User.is_active == True).scalar() or 0,
        "candidates": db.query(func.count(Candidate.id)).scalar() or 0,
        "conversations": db.query(func.count(Conversation.id)).scalar() or 0,
        "messages": db.query(func.count(Message.id)).scalar() or 0,
        "pools": db.query(func.count(TalentPool.id)).scalar() or 0,
        "roles": db.query(func.count(Role.id)).scalar() or 0,
        "analyses_done": db.query(func.count(Conversation.id)).filter(Conversation.analysis_status == "done").scalar() or 0,
        "analyses_pending": db.query(func.count(Conversation.id)).filter(Conversation.analysis_status.in_(("pending", "analyzing"))).scalar() or 0,
        "analyses_failed": db.query(func.count(Conversation.id)).filter(Conversation.analysis_status == "failed").scalar() or 0,
    }


# ========== Endpoints: Settings (API keys etc) ==========

SETTINGS_WHITELIST = {
    # Apify (profile enrichment — replaces the deprecated Proxycurl slot)
    "apify_api_key", "apify_actor_id",
    # Back-compat: old proxycurl key is still accepted but hidden from UI
    "proxycurl_api_key",
    # ATS
    "greenhouse_api_key", "greenhouse_on_behalf_of", "lever_api_key",
    # Misc
    "default_tone",
    # Multi-user / public signup
    "allow_self_register", "auto_approve_signups",
}


@app.get("/api/settings")
def get_settings(
    db: Session = Depends(get_db),
    user: Optional[User] = Depends(current_user),
    _: bool = Depends(require_api_key),
):
    """Return settings state. Never return raw secret values to non-admins — just whether they're set."""
    out = {}
    for key in SETTINGS_WHITELIST:
        val = get_setting(db, key, "")
        is_secret = "api_key" in key or "token" in key
        if is_secret:
            out[key] = {"set": bool(val), "preview": (val[:4] + "•••" + val[-2:]) if val and len(val) > 6 else ""}
        else:
            out[key] = {"set": bool(val), "value": val}
    return out


class SettingIn(BaseModel):
    key: str
    value: str


@app.post("/api/settings")
def save_setting(
    payload: SettingIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    if payload.key not in SETTINGS_WHITELIST:
        raise HTTPException(400, f"Unknown setting key: {payload.key}")
    set_setting(db, payload.key, payload.value)
    return {"ok": True}


# ========== Endpoints: Admin (purge test/mock data) ==========

class PurgeIn(BaseModel):
    confirm: str = ""  # must equal "DELETE" to actually run


@app.post("/api/admin/purge")
def admin_purge(
    payload: PurgeIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    """Wipe recruiting data (candidates, conversations, roles, etc).
    Always keeps: users, sessions, settings.
    """
    if payload.confirm != "DELETE":
        raise HTTPException(400, "Confirmation required: send {confirm: 'DELETE'} to execute.")

    db.query(Task).delete(synchronize_session=False)
    db.query(ReplyClassification).delete(synchronize_session=False)
    db.query(ReplyDraft).delete(synchronize_session=False)
    db.query(CandidateRoleMatch).delete(synchronize_session=False)
    db.query(CampaignRecipient).delete(synchronize_session=False)
    db.query(Campaign).delete(synchronize_session=False)
    db.query(StageHistory).delete(synchronize_session=False)
    db.query(MessageTemplate).delete(synchronize_session=False)
    db.query(Role).delete(synchronize_session=False)
    db.execute(candidate_pool_assoc.delete())
    db.query(TalentPool).delete(synchronize_session=False)
    db.query(Message).delete(synchronize_session=False)
    db.query(Conversation).delete(synchronize_session=False)
    db.query(Candidate).delete(synchronize_session=False)

    db.commit()

    return {
        "ok": True,
        "remaining_candidates": db.query(Candidate).count(),
        "remaining_conversations": db.query(Conversation).count(),
        "remaining_messages": db.query(Message).count(),
    }


# ========== Endpoints: Funnel, stale detector, global search ==========

@app.get("/api/metrics/funnel")
def funnel_metrics(
    role_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    """Sourced → Contacted → Replied → Interested → Shortlisted → Hired."""
    stages_order = ["New", "Engaged", "Sent", "Replied", "Shortlisted", "Follow-up Needed", "Closed"]

    # Base candidate set (all, or via role match)
    if role_id is not None:
        q = db.query(Candidate).join(CandidateRoleMatch, CandidateRoleMatch.candidate_id == Candidate.id)\
              .filter(CandidateRoleMatch.role_id == role_id)
    else:
        q = db.query(Candidate)
    cands = q.all()

    total = len(cands)
    stage_map = {s: 0 for s in stages_order}
    contacted_ids = set()
    replied_ids = set()
    interested_ids = set()
    for c in cands:
        s = c.stage or "New"
        stage_map[s] = stage_map.get(s, 0) + 1
        if c.conversations:
            contacted_ids.add(c.id)
            if any((m.sender or "").strip().lower() != "you" for conv in c.conversations for m in conv.messages):
                replied_ids.add(c.id)
            if any((conv.sentiment or "").lower() == "interested" for conv in c.conversations):
                interested_ids.add(c.id)

    shortlisted = stage_map.get("Shortlisted", 0)
    hired = stage_map.get("Closed", 0)  # treat closed as final — could split later
    contacted = len(contacted_ids)
    replied = len(replied_ids)
    interested = len(interested_ids)

    def pct(n, d): return round(100 * n / d) if d else 0

    return {
        "role_id": role_id,
        "funnel": [
            {"stage": "sourced",     "count": total,       "pct_of_prev": 100},
            {"stage": "contacted",   "count": contacted,   "pct_of_prev": pct(contacted, total)},
            {"stage": "replied",     "count": replied,     "pct_of_prev": pct(replied, contacted)},
            {"stage": "interested",  "count": interested,  "pct_of_prev": pct(interested, replied)},
            {"stage": "shortlisted", "count": shortlisted, "pct_of_prev": pct(shortlisted, interested)},
            {"stage": "hired",       "count": hired,       "pct_of_prev": pct(hired, shortlisted)},
        ],
        "stage_map": stage_map,
        "total_candidates": total,
    }


@app.post("/api/metrics/stale/detect")
def detect_stale(
    days: int = 14,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    """Scan for candidates with no activity in N days and auto-create 'stale' tasks."""
    cutoff = datetime.utcnow() - timedelta(days=days)
    active_stages = ("Engaged", "Sent", "Replied", "Shortlisted", "Follow-up Needed")
    stale = db.query(Candidate).filter(
        Candidate.stage.in_(active_stages),
        or_(Candidate.last_activity_at.is_(None), Candidate.last_activity_at < cutoff),
    ).all()

    created = 0
    for c in stale:
        exists = db.query(Task).filter(
            Task.entity_type == "candidate",
            Task.entity_id == c.id,
            Task.task_type == "stale",
            Task.status == "pending",
        ).first()
        if not exists:
            db.add(Task(
                entity_type="candidate", entity_id=c.id,
                task_type="stale",
                title=f"{c.full_name} has gone quiet ({days}+ days)",
                due_at=datetime.utcnow(),
            ))
            created += 1
    db.commit()
    return {"stale_candidates": len(stale), "tasks_created": created}


@app.get("/api/search")
def global_search(
    q: str = Query("", min_length=1),
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    """Cmd+K search across candidates / roles / pools / templates."""
    if not q:
        return {"candidates": [], "roles": [], "pools": [], "templates": []}
    like = f"%{q}%"
    cands = db.query(Candidate).filter(or_(
        Candidate.full_name.ilike(like),
        Candidate.current_company.ilike(like),
        Candidate.current_title.ilike(like),
        Candidate.email.ilike(like),
    )).limit(10).all()
    roles = db.query(Role).filter(or_(Role.title.ilike(like), Role.company.ilike(like))).limit(10).all()
    pools = db.query(TalentPool).filter(TalentPool.name.ilike(like)).limit(10).all()
    templates = db.query(MessageTemplate).filter(MessageTemplate.name.ilike(like)).limit(10).all()
    return {
        "candidates": [{"id": c.id, "full_name": c.full_name, "subtitle": f"{c.current_title or ''} · {c.current_company or ''}"} for c in cands],
        "roles": [{"id": r.id, "title": r.title, "subtitle": r.company or ""} for r in roles],
        "pools": [{"id": p.id, "name": p.name, "count": len(p.candidates)} for p in pools],
        "templates": [{"id": t.id, "name": t.name, "subtitle": t.template_type or ""} for t in templates],
    }


# ========== Endpoints: Campaigns (bulk outreach) ==========

class CampaignCreateIn(BaseModel):
    name: str
    role_id: Optional[int] = None
    template_id: int
    pool_id: Optional[int] = None
    candidate_ids: Optional[List[int]] = None
    channel: Optional[str] = "linkedin"


@app.post("/api/campaigns")
def create_campaign(
    payload: CampaignCreateIn,
    db: Session = Depends(get_db),
    user: Optional[User] = Depends(current_user),
    _: bool = Depends(require_api_key),
):
    """Create a campaign and auto-draft personalised messages for each recipient."""
    t = db.query(MessageTemplate).filter(MessageTemplate.id == payload.template_id).first()
    if not t:
        raise HTTPException(404, "Template not found")

    recipient_ids: List[int] = []
    if payload.candidate_ids:
        recipient_ids = payload.candidate_ids
    elif payload.pool_id:
        pool = db.query(TalentPool).filter(TalentPool.id == payload.pool_id).first()
        if not pool:
            raise HTTPException(404, "Pool not found")
        recipient_ids = [c.id for c in pool.candidates]
    else:
        raise HTTPException(400, "Provide candidate_ids or pool_id")

    camp = Campaign(
        name=payload.name, role_id=payload.role_id, template_id=payload.template_id,
        pool_id=payload.pool_id, channel=payload.channel or "linkedin",
        status="draft", owner_id=user.id if user else None,
    )
    db.add(camp); db.flush()

    role_dict = _role_dict(t.role) if t.role else {"title": "", "company": "", "must_have": []}

    for cid in recipient_ids:
        c = db.query(Candidate).filter(Candidate.id == cid).first()
        if not c:
            continue
        try:
            variables = {
                "first_name": (c.full_name or "there").split(" ")[0],
                "full_name": c.full_name or "",
                "current_company": c.current_company or "",
                "current_title": c.current_title or "",
                "role_title": role_dict.get("title", ""),
                "role_company": role_dict.get("company", ""),
                "role_location": role_dict.get("location", ""),
                "outreach_angle": role_dict.get("outreach_angle", ""),
                "matched_skill": (role_dict.get("must_have", []) or [""])[0],
            }
            pre_filled = ai_service.render_template(t.body_template, variables)
            draft = ai_service.personalize_outreach(
                role=role_dict,
                candidate=_candidate_summary(c),
                template_body=pre_filled,
                template_type=t.template_type or "first_outreach",
                tone="professional",
            )
        except Exception as e:
            draft = None
            err = str(e)
        else:
            err = None
        db.add(CampaignRecipient(
            campaign_id=camp.id, candidate_id=cid,
            draft_text=draft or "",
            status="draft" if draft else "failed",
            error=err,
        ))
    db.commit()
    return {"id": camp.id, "recipients": len(recipient_ids)}


@app.get("/api/campaigns")
def list_campaigns(
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    rows = db.query(Campaign).order_by(Campaign.updated_at.desc()).all()
    out = []
    for c in rows:
        recs = db.query(CampaignRecipient).filter(CampaignRecipient.campaign_id == c.id).all()
        status_counts = {}
        for r in recs:
            status_counts[r.status] = status_counts.get(r.status, 0) + 1
        out.append({
            "id": c.id, "name": c.name, "status": c.status, "channel": c.channel,
            "role_id": c.role_id, "template_id": c.template_id, "pool_id": c.pool_id,
            "recipient_count": len(recs),
            "status_counts": status_counts,
            "created_at": c.created_at.isoformat() if c.created_at else "",
        })
    return out


@app.get("/api/campaigns/{cid}")
def get_campaign(
    cid: int,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    c = db.query(Campaign).filter(Campaign.id == cid).first()
    if not c:
        raise HTTPException(404, "Campaign not found")
    recs = db.query(CampaignRecipient).filter(CampaignRecipient.campaign_id == cid).all()
    # Resolve candidate names in one query
    cand_ids = [r.candidate_id for r in recs if r.candidate_id]
    cand_map = {x.id: x for x in db.query(Candidate).filter(Candidate.id.in_(cand_ids)).all()}
    return {
        "id": c.id, "name": c.name, "status": c.status, "channel": c.channel,
        "role_id": c.role_id, "template_id": c.template_id, "pool_id": c.pool_id,
        "recipients": [
            {
                "id": r.id,
                "candidate_id": r.candidate_id,
                "candidate_name": (cand_map.get(r.candidate_id).full_name if cand_map.get(r.candidate_id) else ""),
                "current_title": (cand_map.get(r.candidate_id).current_title if cand_map.get(r.candidate_id) else ""),
                "current_company": (cand_map.get(r.candidate_id).current_company if cand_map.get(r.candidate_id) else ""),
                "draft_text": r.draft_text or "",
                "status": r.status,
                "sent_at": r.sent_at.isoformat() if r.sent_at else "",
                "replied_at": r.replied_at.isoformat() if r.replied_at else "",
                "error": r.error or "",
            }
            for r in recs
        ],
    }


class RecipientPatchIn(BaseModel):
    draft_text: Optional[str] = None
    status: Optional[str] = None


@app.patch("/api/campaigns/recipients/{rid}")
def patch_campaign_recipient(
    rid: int, payload: RecipientPatchIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    r = db.query(CampaignRecipient).filter(CampaignRecipient.id == rid).first()
    if not r:
        raise HTTPException(404, "Recipient not found")
    if payload.draft_text is not None:
        r.draft_text = payload.draft_text
    if payload.status is not None:
        r.status = payload.status
        if payload.status == "sent":
            r.sent_at = datetime.utcnow()
    db.commit()
    return {"ok": True}


# CampaignRecipient needs a relation to Candidate for the UI — do it lazily (join-less)
# by resolving candidate_id in the endpoint above.

# ========== Endpoints: ATS push ==========

class AtsPushIn(BaseModel):
    candidate_id: int
    provider: str = "greenhouse"


@app.post("/api/ats/push")
def ats_push(
    payload: AtsPushIn,
    db: Session = Depends(get_db),
    _: bool = Depends(require_api_key),
):
    c = db.query(Candidate).filter(Candidate.id == payload.candidate_id).first()
    if not c:
        raise HTTPException(404, "Candidate not found")
    cand = {
        "full_name": c.full_name,
        "current_title": c.current_title or "",
        "current_company": c.current_company or "",
        "email": c.email or "",
        "phone": c.phone or "",
        "profile_url": c.profile_url or "",
        "tags": c.tags or "",
        "notes": c.notes or "",
        "headline": c.headline or "",
    }
    try:
        if payload.provider == "greenhouse":
            key = get_setting(db, "greenhouse_api_key", "")
            obo = get_setting(db, "greenhouse_on_behalf_of", "")
            if not key or not obo:
                raise HTTPException(400, "Greenhouse not configured. Add API key + user id in Settings.")
            res = ai_service.greenhouse_push_candidate(key, obo, cand)
        elif payload.provider == "lever":
            key = get_setting(db, "lever_api_key", "")
            if not key:
                raise HTTPException(400, "Lever not configured.")
            res = ai_service.lever_push_candidate(key, cand)
        else:
            raise HTTPException(400, f"Unknown ATS provider: {payload.provider}")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(502, f"ATS push failed: {e}")

    # Tag the candidate so the UI knows it was pushed
    existing = (c.tags or "")
    if "pushed_to_ats" not in existing:
        c.tags = (existing + ", pushed_to_ats").strip(", ") if existing else "pushed_to_ats"
        db.commit()
    return {"ok": True, "ats_response": res}


# ========== Endpoints: Excel export ==========

@app.get("/api/export/conversations.xlsx")
def export_conversations_xlsx(
    db: Session = Depends(get_db),
    user: Optional[User] = Depends(current_user),
    _: bool = Depends(require_api_key),
):
    """Generate an Excel workbook with candidates + full conversation transcripts.
    Non-admins only see their own data."""
    wb = Workbook()

    # -- Sheet 1: Candidates overview --
    ws1 = wb.active
    ws1.title = "Candidates"
    headers = [
        "Name", "Company", "Title", "Location", "Stage", "Tags",
        "Sentiment (latest)", "Conversations", "Profile URL", "Last Updated", "Notes"
    ]
    ws1.append(headers)

    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    cand_q = db.query(Candidate)
    if user and not getattr(user, "is_admin", False):
        cand_q = cand_q.filter(Candidate.owner_user_id == user.id)
    candidates = cand_q.order_by(Candidate.updated_at.desc()).all()
    for c in candidates:
        latest_sentiment = ""
        if c.conversations:
            latest_sentiment = c.conversations[-1].sentiment or ""
        ws1.append([
            c.full_name,
            c.current_company or "",
            c.current_title or "",
            c.location or "",
            c.stage or "",
            c.tags or "",
            latest_sentiment,
            len(c.conversations),
            c.profile_url or "",
            c.updated_at.strftime("%Y-%m-%d %H:%M") if c.updated_at else "",
            (c.notes or "")[:500],
        ])

    # Column widths
    widths = [22, 22, 24, 18, 14, 20, 16, 14, 40, 18, 40]
    for i, w in enumerate(widths, start=1):
        ws1.column_dimensions[ws1.cell(row=1, column=i).column_letter].width = w

    # -- Sheet 2: Full conversation transcripts --
    ws2 = wb.create_sheet("Conversations")
    ws2.append(["Candidate", "Captured At", "Sentiment", "Summary", "Sender", "Message"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    for c in candidates:
        for conv in c.conversations:
            for m in conv.messages:
                ws2.append([
                    c.full_name,
                    conv.captured_at.strftime("%Y-%m-%d %H:%M") if conv.captured_at else "",
                    conv.sentiment or "",
                    (conv.summary or "")[:500],
                    m.sender,
                    m.body,
                ])

    widths2 = [22, 18, 14, 50, 18, 80]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w
    # Wrap long messages
    for row in ws2.iter_rows(min_row=2):
        row[-1].alignment = Alignment(wrap_text=True, vertical="top")
        row[3].alignment = Alignment(wrap_text=True, vertical="top")

    # Stream response
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"recruiter_export_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ========== Frontend ==========

@app.get("/api/health")
def health():
    return {"status": "ok"}


# Serve the web UI at / — MUST be mounted LAST so /api/* routes take precedence
if os.path.isdir("static"):
    app.mount("/", StaticFiles(directory="static", html=True), name="static")
